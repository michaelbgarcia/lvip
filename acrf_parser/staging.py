"""Phase 9 - the staging workbook.

The contract between all three lanes of the workflow. Python parses the blank
CRF and pre-fills what history already answers; the agent fills the rows history
could not reach; a human reviews the whole thing in Excel; and it comes back to
Python to be written onto the PDF.

Design follows from who reads it:

* **Copilot / the agent** works best on one flat table with self-describing
  column names, so the work surface is a single sheet with one row per field and
  no merged cells, no nesting.
* **The human** needs to see *why* a row was filled, so the match tier, score and
  source study sit beside the suggestion rather than in a hidden audit log. A
  suggestion that looks as authoritative as a fact is the failure mode here.
* **The importer** needs geometry to place the annotation, but geometry is noise
  to the first two readers and must not be hand-edited. So it lives on a locked
  `Geometry` sheet keyed by `row_id`, out of the way and impossible to typo.

Only `EXACT_KEY` rows arrive as AUTO. Everything else is NEEDS_REVIEW or
NEEDS_MAPPING regardless of score, so no fuzzy match can slip through unlooked-at.

The unit of a row is one **annotation**, not one field. A CRF field routinely
carries several statements - "Date of informed consent" is DSTERM, DSDECOD=
INFORMED CONSENT OBTAINED, RFICDTC and DSSTDTC - so the key is (`row_id`,
`annot_seq`): the field it belongs to, and which of that field's annotations it
is. History exports the whole set it has seen; a reviewer adds the rest by
copying a row and giving it the next free seq. One statement per row is what
lets each annotation keep its own type, colour and placement, and what lets the
importer name the one that is wrong.
"""
from __future__ import annotations

from dataclasses import dataclass, field as dc_field, replace
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import prefill as pf
from .models import FIELD_SCOPE, FORM_SCOPE, Document
from .normalize import statement_key
from .prefill import PrefillIndex, prefill_document, prefill_forms
from .style import HouseStyle, derive_house_style
from .template import RIGHT_OF

SHEET_WORK, SHEET_GEOM = "Annotations", "Geometry"
SHEET_STYLE, SHEET_FORMS, SHEET_README = "HouseStyle", "Forms", "Instructions"
SHEET_FILLS, SHEET_ALTS = "DomainFills", "Alternatives"

STATUSES = ["AUTO", "NEEDS_REVIEW", "NEEDS_MAPPING", "APPROVED", "REJECTED"]

# Where a row's position came from. Recorded on the Geometry sheet because it
# changes what the position *means* - and because the writer needs it: rows that
# share one starting point have to be chained left to right so they do not all
# land on top of each other, and rows that each carry their own recorded point
# must be left exactly where they are.
LEARNED = "history"          # this statement's offset from this field, recorded
LEARNED_SPOT = "history_page"  # its position on the page, when the offset is not usable
HOUSE_STYLE = "house_style"  # the corpus median for annotations of this type
HEADER_BAND = "header_band"  # the page's own header band - no evidence either way

# How far from its field a *learned* offset may put an annotation before it stops
# being evidence and starts being a linking error copied forward. Page fractions.
MAX_LEARNED_OFFSET_X = 0.35
MAX_LEARNED_OFFSET_Y = 0.08

# Reported in `fill_basis` when the fill came from this exact statement's own
# appearance in a previous study rather than from a rule about its type or its
# domain. The most specific evidence there is, and worth saying so.
FILL_SEEN = "this exact statement's own fill in a previous study"

ANNOT_TYPES = ["VARIABLE", "CONSTANT_ASSIGNMENT", "CONDITIONAL_VARIABLE",
               "SUPP_QUALIFIER", "NOT_SUBMITTED", "DERIVATION_RULE",
               "CROSS_REFERENCE", "DOMAIN_HEADER", "NOTE"]
SCOPES = [FIELD_SCOPE, FORM_SCOPE]

# Columns the human and the agent fill in. Everything else is evidence or is locked.
#
# `form_name` is editable, and deliberately so. A blank CRF carries no
# domain-header annotations, so a page with no printed title is inherited by the
# form above it - the Disposition page reads as "Medical History". The reviewer
# is the only one who can see that and fix it, and since the primary key is
# (form_name, field_text), an uncorrectable form name would file the pair under
# the wrong form forever.
#
# `annot_seq` is editable for the same reason: it is how a reviewer says "this
# field needs a second annotation". They copy the row, change the seq, and write
# the second statement - no other column tells the importer that is what they
# meant.
EDITABLE = ("annot_seq", "form_name", "final_variable", "final_annotation",
            "final_annot_type", "status", "reviewer_note", "color_rgb", "fill_rgb",
            "font_name", "font_size", "placement")

HEADERS = [
    ("row_id", 10), ("annot_seq", 10), ("scope", 8), ("form_name", 22), ("page", 6),
    ("field_text", 42),
    ("item_number", 11), ("options", 26), ("control", 9),
    ("suggested_variable", 18), ("suggested_annotation", 26), ("suggested_annot_type", 20),
    ("match_tier", 22), ("match_score", 11), ("match_source", 38), ("match_reason", 52),
    ("known_aliases", 26),
    ("final_variable", 16), ("final_annotation", 26), ("final_annot_type", 18),
    ("status", 15), ("reviewer_note", 30),
    ("color_rgb", 14), ("fill_rgb", 12), ("fill_basis", 46), ("font_name", 11),
    ("font_size", 10), ("placement", 17),
]

FILL_AUTO = PatternFill("solid", fgColor="E8F5E9")      # settled: nothing to do
FILL_REVIEW = PatternFill("solid", fgColor="FFF8E1")    # suggestion: look at it
FILL_MAPPING = PatternFill("solid", fgColor="FFEBEE")   # the agent's real work
FILL_HEADER = PatternFill("solid", fgColor="263238")
FILL_LOCKED = PatternFill("solid", fgColor="F5F5F5")
STATUS_FILL = {"AUTO": FILL_AUTO, "NEEDS_REVIEW": FILL_REVIEW,
               "NEEDS_MAPPING": FILL_MAPPING}


@dataclass
class StagingRow:
    """One annotation of one field of the blank CRF, as it appears in the workbook."""
    row_id: str
    values: dict[str, Any]
    geometry: dict[str, Any]
    annot_seq: int = 1
    # Rival answers history had for this row's entity - another study's wording
    # of the same markup, or a weaker tier that was not taken. Never rows on the
    # work sheet: a row there is a promise that something gets drawn. They go to
    # the 'Alternatives' sheet as evidence, for the agent to weigh against
    # context this pipeline does not have. Carried on annot_seq 1 only, because
    # an alternative is rival to the whole set, not to one statement in it.
    alternates: list[Any] = dc_field(default_factory=list)

    @property
    def key(self) -> tuple[str, int]:
        """What identifies this row in the workbook, and on the way back in."""
        return (self.row_id, self.annot_seq)


def build_staging(doc: Document, index: PrefillIndex | None = None,
                  house: HouseStyle | None = None) -> list[StagingRow]:
    """Pre-fill a parsed blank CRF into staging rows.

    One row per annotation history proposes, and always at least one per field:
    a field history has nothing for still gets its NEEDS_MAPPING row, because a
    field with no row is a field nobody is asked about.

    The same reasoning is what puts the **form-level** rows here. A page's domain
    headers and form-level constants belong to no field, so before there was a
    scope column there was no row they could occupy - and a layer with no rows is
    a layer nobody is asked about, which is precisely why annotated CRFs came out
    of this pipeline missing the markup across the top of every page. Each page
    that belongs to a form now leads with its own rows, in reading order: the
    form's markup first, then the fields under it.

    With no index the sheet is still produced - every row simply arrives as
    NEEDS_MAPPING, which is the honest state of a first study with no history.
    """
    index = index or PrefillIndex()
    house = house or derive_house_style([])
    results = {r.field_id: r for r in prefill_document(doc, index)}
    form_results = {r.field_id: r for r in prefill_forms(doc, index)}

    rows: list[StagingRow] = []
    for page in doc.pages:
        if page.anchor is not None:
            r = form_results.get(page.anchor.id)
            for seq, candidate in enumerate(r.annotations if r else [None], start=1):
                row = _form_row(page, page.anchor, r, candidate, seq, house)
                _attach_alternates(row, r if seq == 1 else None)
                rows.append(row)
        carriers = _carriers(page, results)
        for fld in page.fields:
            r = results.get(fld.id)
            candidates = (r.annotations if r else [None]) if carriers.get(fld.id, True) else [None]
            for seq, candidate in enumerate(candidates, start=1):
                row = _row(doc, page, fld, r, candidate, seq, house)
                _attach_alternates(row, r if seq == 1 else None)
                rows.append(row)
    return rows


def _attach_alternates(row: StagingRow, result) -> None:
    """Hang the row's rival answers on it, and say on the row that they exist.

    Two kinds arrive as `Prefill.alternates` and both belong here: another study
    worded this markup differently, and a weaker tier reached the same field.
    Both are answers history had and the rule did not take.

    Filtered, because an alternative that is not an alternative is noise on a
    sheet whose whole value is that every line on it is a real decision. Two
    tiers reaching `BRTHDTC` by different routes is one answer found twice, not
    a choice - `statement_key` is what says so, the same word-set identity used
    everywhere else. The same statement offered by two losing studies collapses
    for the same reason.

    The pointer goes in `match_reason` because that is the column a reviewer and
    the agent already read for *why*, and the honest answer to "why this one" is
    partly "because others were passed over". A row with no alternatives says
    nothing, so the note means something when it appears.
    """
    if result is None or not result.alternates:
        return
    chosen = statement_key(row.values.get("suggested_annotation") or "")
    seen, keep = {chosen}, []
    for alt in result.alternates:
        key = statement_key(alt.annotation_text or "")
        if not key or key in seen:
            continue
        seen.add(key)
        keep.append(alt)
    if not keep:
        return
    row.alternates = keep
    reason = row.values.get("match_reason") or ""
    note = (f"{len(keep)} alternative{'' if len(keep) == 1 else 's'} in history "
            f"- see the '{SHEET_ALTS}' sheet")
    row.values["match_reason"] = f"{reason}; {note}" if reason else note


def _carriers(page, results) -> dict[str, bool]:
    """Which of a page's fields should actually carry the markup for their key.

    A log form repeats one question down the page - twelve blank medication rows,
    seventeen adverse-event rows - and every repetition normalizes to the same
    `(form, field_text)` key. Pre-fill answers each of them, which is right: they
    are all that field. But an annotator does not write `CMTRT` beside all twelve
    rows, they write it once against the column, and a pipeline that draws it
    twelve times has not been thorough, it has defaced the page.

    Which one keeps it is not a guess either. History recorded where the
    statement was drawn, so the occurrence that keeps it is the one nearest that
    spot - and where a key was drawn on several occurrences, that many keep it.
    The rest get their usual NEEDS_MAPPING row: a reviewer is still asked about
    every field, they are simply not handed the same answer twelve times.

    Fields whose key appears once on the page - almost all of them - are
    untouched, and so is any key history has no position for.
    """
    by_key: dict[tuple[str, str], list] = {}
    for fld in page.fields:
        by_key.setdefault(fld.key, []).append(fld)

    out: dict[str, bool] = {}
    for key, fields in by_key.items():
        if len(fields) < 2:
            continue
        spots = [c.anchor for f in fields[:1]
                 for c in (results[f.id].annotations if results.get(f.id) else [])
                 if c is not None and c.anchor]
        if not spots:
            continue                      # no recorded position: change nothing
        w = page.width or 1.0
        h = page.height or 1.0
        chosen = set()
        for spot in spots:
            x, y = spot["rel_x_pct"] * w, spot["rel_y_pct"] * h
            near = min(fields, key=lambda f: (f.bbox.cx - x) ** 2 + (f.bbox.cy - y) ** 2)
            chosen.add(near.id)
        for f in fields:
            out[f.id] = f.id in chosen
    return out


def _form_row(page, anchor, result, best, seq: int, house: HouseStyle) -> StagingRow:
    """One form-level annotation, placed where history drew it.

    Form-level markup is the case with no field to be positioned against, so the
    house style has nothing to say about where it goes - its offsets are measured
    from *field* markup ("12pt right of the label"), and an anchor is not a
    label. Two answers, in order:

    * **Where this study drew this exact statement**, as a fraction of the page.
      A real aCRF does not stack its domain headers in one band: the MSG
      Demography page carries `DM=Demographics` at the top, `SC=Subject
      Characteristics` beside the family-status block halfway down, and
      `DS=Disposition` above the consent date at the foot. Redrawing all three at
      the top left is not a small error - it is most of a page's height.
    * **The page's header band**, for a statement history has no position for.
      The first lands on the anchor and the rest chain rightwards off each other,
      which is how a row of domain headers is drawn and read.

    `anchor_source` records which of the two answered, so the writer knows
    whether these rows share a starting point (chain them) or each carry their
    own (leave them where they are), and a reviewer can see it in the sheet.
    """
    status = (result.status_of(best) if result and best else pf.NEEDS_MAPPING_STATUS)
    rule, basis = _rule_for(house, best)
    learned = best.anchor if best else None
    return StagingRow(
        row_id=anchor.id,
        annot_seq=seq,
        values={
            "row_id": anchor.id,
            "annot_seq": seq,
            "scope": FORM_SCOPE,
            "form_name": anchor.form_name,
            "page": anchor.page,
            "field_text": anchor.text,
            "item_number": "", "options": "", "control": "",
            **_suggestion_values(result, best, status),
            "color_rgb": _hex(rule.text_color),
            "fill_rgb": _hex(rule.fill_color),
            "fill_basis": basis,
            "font_name": rule.font_name,
            "font_size": rule.font_size or "",
            "placement": RIGHT_OF,
        },
        geometry={
            "row_id": anchor.id, "annot_seq": seq, "page": anchor.page,
            "page_width": page.width, "page_height": page.height,
            **(learned or anchor.bbox.relative(page.width, page.height)),
            "offset_x_pct": 0.0, "offset_y_pct": 0.0,
            "box_w_pct": best.box_width if best else 0.0,
            "anchor_source": LEARNED if learned else HEADER_BAND,
            "group_id": "", "field_confidence": "",
        },
    )


def _rule_for(house: HouseStyle, best) -> tuple[Any, str]:
    """Style for one proposed annotation, plus where its fill came from.

    The house style answers first, then the statement's own recorded appearance
    overrides it where history has one. Both are measurements, so this is not a
    guess beating a rule - it is the more specific measurement beating the more
    general one, the same order `_placement_for` puts them in.

    The house style still decides for everything history has never seen, and the
    HouseStyle sheet still reports the corpus conventions and their agreement,
    which is what a reviewer needs in order to settle a convention *once*.

    Fill follows the same order, and `fill_basis` says which answered. The
    domain rule is still what reaches a statement the corpus has never seen -
    that is the whole point of measuring fill per domain - but where the corpus
    drew *this* statement, that colour wins. It has to: a domain rule is a mode
    over a whole corpus, and on the MSG CRF, where 188 of 206 boxes are cyan, it
    reports cyan for DS markup that the sponsor drew yellow on the very page the
    distinction matters.
    """
    annot_type = best.annot_type if best and best.annot_type else "VARIABLE"
    text = best.annotation_text if best else ""
    rule, basis = house.for_annotation(annot_type, text)
    seen = best.style if best else {}
    overrides = {k: v for k, v in seen.items()
                 if k in ("text_color", "font_name", "font_size") and v}
    if seen.get("fill_color") and seen["fill_color"] != rule.fill_color:
        overrides["fill_color"] = seen["fill_color"]
        basis = FILL_SEEN
    return (replace(rule, **overrides) if overrides else rule), basis


def _suggestion_values(result, best, status: str) -> dict[str, Any]:
    """The evidence and decision columns, shared by field and form rows."""
    return {
        "suggested_variable": best.variable if best else "",
        "suggested_annotation": best.annotation_text if best else "",
        "suggested_annot_type": best.annot_type if best else "",
        "match_tier": best.tier if best else pf.NEEDS_MAPPING,
        "match_score": round(best.confidence, 3) if best else 0.0,
        "match_source": best.source if best else "",
        "match_reason": "; ".join(best.evidence) if best else "",
        "known_aliases": ", ".join(result.aliases) if result else "",
        # An AUTO row is pre-accepted; anything else is left blank on purpose,
        # so a reviewer cannot mistake a suggestion for a decision.
        "final_variable": best.variable if status == pf.AUTO else "",
        "final_annotation": best.annotation_text if status == pf.AUTO else "",
        "final_annot_type": best.annot_type if status == pf.AUTO else "",
        "status": status,
        "reviewer_note": "",
    }


def _row(doc, page, fld, result, best, seq: int, house: HouseStyle) -> StagingRow:
    """One annotation of one field, positioned from the best evidence available.

    Three answers, in order of how specific the evidence is:

    1. **The offset history recorded from this field.** Portable - it survives
       the form being re-flowed - and it is a claim about this exact statement.
    2. **The page position history recorded**, where that offset is missing or
       implausible. Less portable, because it says nothing about the field it
       belongs to, but it is still where somebody put this statement rather than
       an average of where they put statements like it. That happens on log
       forms, where one key names seventeen repeating rows and the offset the
       linker measured is to whichever one it picked.
    3. **The house style**, the median over every annotation of this type in the
       corpus - the right answer for a statement nobody has seen.

    `anchor_source` records which, so a reviewer asking why a box landed
    somewhere gets the actual reason rather than a plausible one.
    """
    status = (result.status_of(best) if result and best
              else pf.NEEDS_MAPPING_STATUS)
    rule, basis = _rule_for(house, best)
    placement, off_x, off_y, source = _placement_for(rule, best)
    spot = None if source == LEARNED else _spot_geometry(best, page)
    if spot:
        placement, off_x, off_y, source = RIGHT_OF, 0.0, 0.0, LEARNED_SPOT
    return StagingRow(
        row_id=fld.id,
        annot_seq=seq,
        values={
            "row_id": fld.id,
            "annot_seq": seq,
            "scope": FIELD_SCOPE,
            "form_name": fld.form_name,
            "page": fld.page,
            "field_text": fld.text,
            "item_number": fld.item_number,
            "options": " | ".join(fld.option_texts),
            "control": ",".join(fld.control_kinds),
            **_suggestion_values(result, best, status),
            "color_rgb": _hex(rule.text_color),
            # Blank when the corpus draws no background: an unfilled box is a
            # real house style, not a missing value to be invented. Blank *also*
            # where the corpus colour-codes by domain and this statement's domain
            # could not be resolved - `fill_basis` says which of the two it is.
            "fill_rgb": _hex(rule.fill_color),
            "fill_basis": basis,
            "font_name": rule.font_name,
            "font_size": rule.font_size or "",
            "placement": placement,
        },
        geometry={
            "row_id": fld.id, "annot_seq": seq, "page": fld.page,
            "page_width": page.width, "page_height": page.height,
            **(spot or fld.bbox.relative(page.width, page.height)),
            "offset_x_pct": off_x, "offset_y_pct": off_y,
            "box_w_pct": best.box_width if best else 0.0,
            "anchor_source": source,
            "group_id": fld.group_id, "field_confidence": fld.confidence,
        },
    )


def _spot_geometry(best, page) -> dict[str, float] | None:
    """This statement's own recorded position, as a zero-width point to start at.

    Used in place of the field's box when the offset from that field is not
    usable. A zero-width box with `right_of_field` and no offset reproduces the
    recorded left edge exactly, and lets the new box take whatever width its own
    text needs - the same shape `_form_row` uses, for the same reason.
    """
    return best.anchor if best else None


def _placement_for(rule, best) -> tuple[str, float, float, str]:
    """(placement, offset_x_pct, offset_y_pct, where it came from) for one row.

    History first, house style second. The distinction is worth keeping visible:
    a row placed from the corpus is a claim about *this* statement on *this*
    field, and a row placed from the house style is a claim about annotations of
    this type in general. When a reviewer asks why a box landed somewhere, those
    are two different answers.

    History is only believed within `MAX_LEARNED_OFFSET`. An offset is the gap
    between an annotation and the field it was *linked to*, so an implausible one
    is not a fact about how this sponsor draws markup - it is a linking mistake,
    and using it would copy that mistake onto the new CRF at full confidence.
    Real aCRF markup sits beside its label: on the MSG corpus the offsets run to
    about a third of the page width, and the handful beyond that are all markup
    the linker attached to a field on the other side of a log-form grid.
    """
    drawn = best.drawn if best else {}
    if drawn.get("relative_label"):
        dx = round(float(drawn.get("offset_x_pct") or 0.0), 4)
        dy = round(float(drawn.get("offset_y_pct") or 0.0), 4)
        if abs(dx) <= MAX_LEARNED_OFFSET_X and abs(dy) <= MAX_LEARNED_OFFSET_Y:
            return str(drawn["relative_label"]), dx, dy, LEARNED
    return rule.placement, rule.offset_x_pct, rule.offset_y_pct, HOUSE_STYLE


def _hex(color) -> str:
    """RGB floats to the hex string a spreadsheet can show and a human can read."""
    if not color:
        return ""
    return "#" + "".join(f"{max(0, min(255, round(c * 255))):02X}" for c in color[:3])


# --- workbook --------------------------------------------------------------
def write_staging(doc: Document, path: str | Path, index: PrefillIndex | None = None,
                  house: HouseStyle | None = None) -> Path:
    """Write the staging workbook for one blank CRF."""
    rows = build_staging(doc, index, house)
    house = house or derive_house_style([])
    wb = Workbook()
    _work_sheet(wb, rows)
    _geometry_sheet(wb, rows)
    _style_sheet(wb, house)
    _fills_sheet(wb, house)
    _alternatives_sheet(wb, rows)
    _forms_sheet(wb, doc, index or PrefillIndex())
    _readme_sheet(wb, doc, rows)
    wb.remove(wb["Sheet"]) if "Sheet" in wb.sheetnames else None
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _work_sheet(wb: Workbook, rows: list[StagingRow]) -> None:
    ws = wb.create_sheet(SHEET_WORK, 0)
    names = [h for h, _ in HEADERS]
    ws.append(names)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for r in rows:
        ws.append([r.values.get(n, "") for n in names])

    editable = {names.index(n) + 1 for n in EDITABLE}
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        status = ws.cell(row=i, column=names.index("status") + 1).value
        fill = STATUS_FILL.get(status)
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in
                                       {names.index("field_text") + 1,
                                        names.index("match_reason") + 1})
            # Lock everything the importer relies on; leave the work columns open.
            cell.protection = Protection(locked=cell.column not in editable)
            if cell.column not in editable:
                cell.fill = FILL_LOCKED
        if fill:
            ws.cell(row=i, column=names.index("status") + 1).fill = fill

    _validate(ws, names, "status", STATUSES, len(rows))
    _validate(ws, names, "final_annot_type", ANNOT_TYPES, len(rows))
    _validate(ws, names, "scope", SCOPES, len(rows))
    for idx, (_, width) in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions
    # Sheet protection is deliberately not enabled: it would block Copilot from
    # writing at all. The locked flags mark intent, and the importer is what
    # actually enforces it - validation on the way back in, not a UI lock.


def _validate(ws, names: list[str], column: str, options: list[str], n_rows: int) -> None:
    col = get_column_letter(names.index(column) + 1)
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{max(2, n_rows + 1)}")


def _geometry_sheet(wb: Workbook, rows: list[StagingRow]) -> None:
    """Locked, keyed by (row_id, annot_seq). The importer's input, not the human's.

    A sibling row a reviewer adds has no geometry row of its own - they cannot
    write to this sheet, and should not have to. The importer inherits the
    field's geometry from seq 1 instead; every row of one field describes the
    same box on the page, so there is nothing to lose in doing so.
    """
    ws = wb.create_sheet(SHEET_GEOM)
    keys = ["row_id", "annot_seq", "page", "page_width", "page_height",
            "rel_x_pct", "rel_y_pct", "rel_w_pct", "rel_h_pct",
            "offset_x_pct", "offset_y_pct", "box_w_pct", "anchor_source",
            "group_id", "field_confidence"]
    ws.append(keys)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
    for r in rows:
        ws.append([r.geometry.get(k, "") for k in keys])
    for i in range(1, len(keys) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "B2"


ALT_HEADERS = [
    ("row_id", 10), ("scope", 8), ("form_name", 22), ("page", 6),
    ("field_text", 34), ("chosen_annotation", 26), ("chosen_source", 30),
    ("alt_annotation", 26), ("alt_variable", 16), ("alt_annot_type", 20),
    ("alt_source", 30), ("alt_tier", 22), ("alt_score", 10), ("alt_trust", 16),
    ("why_not_chosen", 56),
]


def _alternatives_sheet(wb: Workbook, rows: list[StagingRow]) -> None:
    """Rival answers history had, and did not get. Evidence, never rows.

    The deterministic rule has to pick one answer per entity, and a corpus of
    several studies routinely holds more than one: two sponsors word the same
    mapping differently, or a fuzzy tier reached the field after an exact one
    did. Merging them onto the page is what put the same domain header on it
    three times, so only one is drawn - but the losers are not wrong, they are
    unchosen, and discarding them makes the corpus look unanimous when it was
    not.

    They cannot live on the 'Annotations' sheet. A row there is a promise: the
    importer draws every approved row, so a rival wording sitting next to its
    default is one careless APPROVE away from both being on the page - and a
    human skimming a sheet where half the rows are alternatives cannot see the
    work. So they get their own sheet, locked, keyed by `row_id` back to the row
    they are rival to.

    This is the handoff. Choosing between wordings is a semantic question about
    a CRF this pipeline has never seen, which is the agent's job and not a
    tie-break's. The rule picks a defensible default so nothing is ever blank;
    the agent reads this sheet, and where it knows better it edits
    `final_annotation` on the work sheet. The 'Instructions' sheet says so.

    `chosen_annotation` is the row the alternative is filed against, which on a
    FORM row is the *first* statement of the page's markup. The alternative is
    rival to that whole set - one study's page against another's - not to that
    one statement, because a page's markup is composed together or not at all.
    """
    ws = wb.create_sheet(SHEET_ALTS)
    names = [h for h, _ in ALT_HEADERS]
    ws.append(names)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        for alt in r.alternates:
            ws.append([
                r.row_id, r.values.get("scope", ""), r.values.get("form_name", ""),
                r.values.get("page", ""), r.values.get("field_text", ""),
                r.values.get("suggested_annotation", ""),
                r.values.get("match_source", ""),
                alt.annotation_text, alt.variable, alt.annot_type,
                alt.source, alt.tier, round(alt.confidence, 3), alt.trust,
                "; ".join(alt.evidence),
            ])
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.protection = Protection(locked=True)
            cell.fill = FILL_LOCKED
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in
                                       {names.index("field_text") + 1,
                                        names.index("why_not_chosen") + 1})
    for i, (_, width) in enumerate(ALT_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "B2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def _style_sheet(wb: Workbook, house: HouseStyle) -> None:
    """The derived house style, with agreement, so unsettled rules get decided once."""
    ws = wb.create_sheet(SHEET_STYLE)
    ws.append(["scope", "samples", "color_rgb", "color_agreement",
               "fill_rgb", "fill_agreement", "fill_samples", "font_name",
               "font_size", "size_agreement", "placement", "placement_agreement",
               "offset_x_pct", "offset_y_pct", "settled", "evidence"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
    for rule in [house.default, *[house.by_type[k] for k in sorted(house.by_type)]]:
        ws.append([rule.scope, rule.samples, _hex(rule.text_color), rule.color_agreement,
                   _hex(rule.fill_color), rule.fill_agreement, rule.fill_samples,
                   rule.font_name, rule.font_size, rule.size_agreement, rule.placement,
                   rule.placement_agreement, rule.offset_x_pct, rule.offset_y_pct,
                   "yes" if rule.settled else "NO", "; ".join(rule.evidence)])
        if not rule.settled:
            for cell in ws[ws.max_row]:
                cell.fill = FILL_REVIEW
    for i, w in enumerate([22, 9, 11, 15, 11, 13, 12, 11, 10, 14, 16, 18, 13, 13, 9, 70],
                          start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _fills_sheet(wb: Workbook, house: HouseStyle) -> None:
    """The study's fill palette: which colour goes with which domain.

    Its own sheet because "which domain is which colour?" is a question a
    reviewer asks as a whole, and the answer is unreadable spread across a
    thousand rows of the work sheet. Statement rows are the exceptions: markup
    whose domain cannot be read off the text (DM's own variables carry no
    prefix), listed with the fill the corpus actually drew it in.
    """
    ws = wb.create_sheet(SHEET_FILLS)
    ws.append(["scope", "fill_rgb", "fill_agreement", "fill_samples", "samples",
               "text_color", "evidence"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
    note = ("fill varies by domain in this corpus - a row whose domain cannot be "
            "resolved is left blank for you to decide"
            if house.fill_varies_by_domain else
            "fill does not vary by domain in this corpus - every row takes the "
            "house style fill and this sheet is reference only")
    for rule in house.fill_rules():
        ws.append([rule.scope, _hex(rule.fill_color), rule.fill_agreement,
                   rule.fill_samples, rule.samples, _hex(rule.text_color),
                   "; ".join(rule.evidence)])
        if rule.fill_samples and rule.fill_agreement < 0.7:
            for cell in ws[ws.max_row]:
                cell.fill = FILL_REVIEW
    ws.append([])
    ws.append([note])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    for i, w in enumerate([34, 11, 15, 13, 9, 12, 70], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _forms_sheet(wb: Workbook, doc: Document, index: PrefillIndex) -> None:
    """Form inventory, with the domain resolved.

    A blank CRF has no domain-header annotations, so `domain` is empty on every
    form it parses. Where history knows the form, its domain is filled in from
    there and `domain_source` says so - the importer validates against this, and
    a domain nobody could establish must not read as an agreed one.
    """
    ws = wb.create_sheet(SHEET_FORMS)
    ws.append(["form_name", "domain", "domain_source", "pages", "continuation_pages",
               "confidence", "source", "evidence"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
    for f in doc.forms:
        domain, origin = f.domain, "this CRF"
        if not domain:
            domain, origin = index.domain_for(f.name), "previous studies"
        ws.append([f.name, domain, origin if domain else "unknown",
                   ", ".join(map(str, f.pages)),
                   ", ".join(map(str, f.continuation_pages)), round(f.confidence, 2),
                   f.source, "; ".join(f.evidence)])
    for i, w in enumerate([26, 9, 17, 14, 18, 11, 22, 60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _readme_sheet(wb: Workbook, doc: Document, rows: list[StagingRow]) -> None:
    """Written for both readers: a human skimming, and an agent given the sheet."""
    ws = wb.create_sheet(SHEET_README)
    counts: dict[str, int] = {}
    for r in rows:
        counts[r.values["status"]] = counts.get(r.values["status"], 0) + 1
    fields = len({r.row_id for r in rows})
    lines = [
        ("aCRF staging workbook", True),
        (f"Source: {Path(doc.path).name} · {doc.page_count} pages · "
         f"{len(doc.forms)} forms · {fields} fields · {len(rows)} annotation rows", False),
        (f"Generated: {datetime.now(timezone.utc).isoformat(timespec='seconds')}", False),
        ("", False),
        ("How to use this workbook", True),
        ("1. Work the 'Annotations' sheet. One row per ANNOTATION, not per field:", False),
        ("   a field with four annotations has four rows, sharing one row_id and", False),
        ("   numbered 1, 2, 3, 4 in annot_seq.", False),
        ("1b. Rows where scope = FORM are the markup across the TOP of a page -", False),
        ("   the domain header(s) and any form-level constants. They belong to the", False),
        ("   form, not to a question, so field_text reads '[form header] <form>'.", False),
        ("   A page usually carries more than one: two domain headers side by side", False),
        ("   (DS=Disposition, DM=Demographics) plus constants such as", False),
        ("   'DSCAT = PROTOCOL MILESTONE'. Add a row per statement, same way as", False),
        ("   for a field: copy the row and bump annot_seq. They are drawn left to", False),
        ("   right across the header band in annot_seq order.", False),
        ("2. Rows are pre-filled from previously annotated studies where the", False),
        ("   (form_name, field_text) pair was seen before. Those arrive as AUTO", False),
        ("   with final_variable already set - confirm, do not retype.", False),
        ("3. NEEDS_REVIEW means a suggestion was found by similarity, not by an", False),
        ("   exact match. Check match_tier, match_score and match_source, then", False),
        ("   put your decision in final_variable / final_annotation.", False),
        ("4. NEEDS_MAPPING means history had nothing. These need real mapping work.", False),
        ("5. Check form_name. A page with no printed title is assumed to belong to", False),
        ("   the form above it, which is often wrong on a blank CRF. Correct it here -", False),
        ("   the mapping is stored under (form_name, field_text), so a wrong form name", False),
        ("   files the answer where nobody will find it again.", False),
        ("6. Set status to APPROVED or REJECTED on every row before importing.", False),
        ("", False),
        ("Reviewing the 'Alternatives' sheet", True),
        ("- Where several past studies annotated the same thing differently, only", False),
        ("  one answer is on the 'Annotations' sheet. It was chosen by rule -", False),
        ("  strongest trust, then score, then the fullest record - and the rule", False),
        ("  knows nothing about THIS study. The wordings it passed over are on the", False),
        ("  'Alternatives' sheet, keyed by row_id to the row they are rival to.", False),
        ("- On a FORM row the alternatives are rival to that PAGE'S WHOLE SET of", False),
        ("  markup, not to the single statement in chosen_annotation: one study's", False),
        ("  header-plus-constants against another's. Weigh the set, not the line.", False),
        ("- Read that sheet. For each row_id on it, compare alt_annotation against", False),
        ("  the chosen_annotation and decide, using context the rule does not have:", False),
        ("  this study's protocol and SDTM conventions, the domain, sibling fields", False),
        ("  on the same form, and which past study the alt_source names. A rival", False),
        ("  from a study in the same therapeutic area usually beats a default that", False),
        ("  won on a tie-break.", False),
        ("- To take an alternative, edit final_annotation (and final_variable /", False),
        ("  final_annot_type) on the matching 'Annotations' row, and say why in", False),
        ("  reviewer_note. Do NOT add a row for it: the alternative REPLACES the", False),
        ("  default, it is not a second annotation. Adding a row draws both, which", False),
        ("  is the exact fault this sheet exists to prevent.", False),
        ("- Leaving the default is a valid outcome and needs no edit. Alternatives", False),
        ("  are unchosen, not wrong; most rows should keep what they arrived with.", False),
        ("- match_reason on a row says 'see the Alternatives sheet' when that row", False),
        ("  has any. A row without that note had exactly one answer in history.", False),
        ("", False),
        ("Adding a second annotation to a field", True),
        ("- Copy the whole row, paste it directly beneath, and put the next free", False),
        ("  number in annot_seq (leave annot_seq blank and it will be numbered for", False),
        ("  you, with a warning). Keep row_id exactly as it is - that is what says", False),
        ("  which field the annotation belongs to.", False),
        ("- Write one statement per row. 'DSTERM' and 'RFICDTC' are two rows, not", False),
        ("  one cell holding both: each gets its own type, colour and placement,", False),
        ("  and they are drawn side by side in annot_seq order.", False),
        ("- Delete a surplus row rather than blanking it, but never delete every", False),
        ("  row of a field - each field needs at least one.", False),
        ("", False),
        ("Rules", True),
        ("- The 'Alternatives' sheet is evidence and is regenerated on export.", False),
        ("  Edits to it are ignored on import; act on it by editing 'Annotations'.", False),
        ("- Only fill the unshaded columns. Shaded columns are evidence and are", False),
        ("  regenerated on every export; edits to them are ignored on import.", False),
        ("- Do not sort or reorder rows, and do not invent a row_id. row_id is the", False),
        ("  key that ties each row to its position on the PDF; a row_id that is not", False),
        ("  in this CRF cannot be placed. Adding rows is allowed only as described", False),
        ("  above: a copy of an existing row, with a new annot_seq.", False),
        ("- (row_id, annot_seq) must be unique. Two rows with the same pair is an", False),
        ("  error, because neither can be told from the other on the way back in.", False),
        ("- Formatting columns are pre-filled from the house style measured across", False),
        ("  previous studies. Change them only for a deliberate exception.", False),
        ("- See the 'HouseStyle' sheet for rules marked settled=NO. Those are cases", False),
        ("  where past studies disagreed and someone needs to decide once.", False),
        ("- fill_rgb may be blank for two different reasons, and fill_basis says", False),
        ("  which: either the study draws unfilled boxes, or the study colour-codes", False),
        ("  by domain and this statement's domain could not be read off its text.", False),
        ("  The second needs a decision - see the 'DomainFills' sheet for the", False),
        ("  palette and put the right colour in. DM's own variables are the usual", False),
        ("  case: nothing in 'RFICDTC' says DM, so it cannot be resolved for you.", False),
        ("", False),
        ("Status of this workbook", True),
    ]
    for status in STATUSES:
        if counts.get(status):
            lines.append((f"  {status}: {counts[status]} row(s)", False))
    for text, bold in lines:
        ws.append([text])
        if bold:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 96


def summarize_staging(rows: list[StagingRow]) -> dict[str, Any]:
    tiers: dict[str, int] = {}
    statuses: dict[str, int] = {}
    for r in rows:
        tiers[r.values["match_tier"]] = tiers.get(r.values["match_tier"], 0) + 1
        statuses[r.values["status"]] = statuses.get(r.values["status"], 0) + 1
    total = len(rows) or 1
    per_field: dict[str, int] = {}
    for r in rows:
        per_field[r.row_id] = per_field.get(r.row_id, 0) + 1
    form_rows = [r for r in rows if r.values.get("scope") == FORM_SCOPE]
    field_rows = [r for r in rows if r.values.get("scope") != FORM_SCOPE]
    return {"rows": len(rows), "fields": len({r.row_id for r in field_rows}),
            "alternatives": sum(len(r.alternates) for r in rows),
            "rows_with_alternatives": sum(1 for r in rows if r.alternates),
            "by_tier": dict(sorted(tiers.items())),
            "by_status": dict(sorted(statuses.items())),
            "auto_fill_rate": round(statuses.get("AUTO", 0) / total, 3),
            "form_rows": len(form_rows),
            "forms_with_several": sum(
                1 for rid in {r.row_id for r in form_rows} if per_field[rid] > 1),
            "undecided_fills": sum(1 for r in rows if not r.values.get("fill_rgb")
                                   and r.values.get("fill_basis")),
            "multi_annotation_fields": len(
                {r.row_id for r in field_rows if per_field[r.row_id] > 1})}
