"""Phase 9c - reading the staging workbook back.

The return leg of the contract. The workbook went out to Excel, where a human
and an agent edited it, and comes back to be written onto the PDF. Everything
that could have gone wrong in between happened in a spreadsheet, unversioned, by
people who could sort a column without thinking about it.

So validation is **per row, not per workbook**. A single unparseable colour in
row 200 must not reject the other 199 - the reviewer would fix it, re-import,
and meet the next problem one round trip later. Each row carries its own
verdict, the good ones import, and the bad ones come back naming their own
problem so they can all be fixed in one pass.

Severity is split for the same reason. `ERROR` blocks the row: without a
resolvable colour or geometry there is nothing to draw. `WARNING` lets the row
through flagged: a variable whose prefix disagrees with the form's domain is
usually wrong, but DM's own variables carry no prefix at all and identifiers
belong to every domain, so refusing them would be worse than noting them.

Two failures are worth calling out because they are silent otherwise:

* **A sorted or deleted row.** `row_id` is the only thing tying a spreadsheet
  row to a position on the PDF. Rows are matched by it, never by position, and
  a missing or unknown one is an error rather than an off-by-one annotation.
* **An unevaluated formula.** If the agent writes `=CONCAT(...)` and the file is
  never opened in Excel, the cached value is empty and the cell silently reads
  as blank. That is detected and reported rather than imported as "no decision".

Several rows per field
----------------------
A row is one annotation, and a field may have several - so the key is
(`row_id`, `annot_seq`) and *added rows are legitimate*. That is the one place
this module deliberately trusts the sheet: a row whose `row_id` names a real
field is accepted however it got there, because copying a row and bumping the
seq is exactly how a reviewer says "this field needs a second annotation".

What is still guarded is everything that could make an added row ambiguous. A
duplicated (row_id, seq) is an error - two rows nothing can tell apart. A blank
seq is numbered automatically and flagged, because "next free" is unambiguous
but silently choosing for the reviewer is not. A `row_id` that is not a field of
this CRF is still an error. And a field with *no* rows left is still an error:
deleting a row is how an annotation is removed, deleting all of them is how a
field goes unannotated by accident.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field as dc_field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from . import annotations as ann
from .models import FIELD_SCOPE, FORM_SCOPE, Document
from .normalize import normalize, statement_key
from .staging import (ANNOT_TYPES, SCOPES, SHEET_FORMS, SHEET_GEOM, SHEET_WORK,
                      STATUSES)
from .template import ABOVE, BELOW, LEFT_OF, OVERLAPS, RIGHT_OF

ERROR, WARNING = "ERROR", "WARNING"
APPROVED, REJECTED = "APPROVED", "REJECTED"

PLACEMENTS = {RIGHT_OF, LEFT_OF, ABOVE, BELOW, OVERLAPS}
MIN_FONT_PT, MAX_FONT_PT = 4.0, 24.0
GEOM_KEYS = ("page_width", "page_height", "rel_x_pct", "rel_y_pct",
             "rel_w_pct", "rel_h_pct")

_HEX = re.compile(r"^#?([0-9A-Fa-f]{6})$")
_VARIABLE = re.compile(r"^(?:[A-Z]{2}\.)?[A-Z][A-Z0-9]{1,7}$")
_SUFFIX = re.compile(r"^[A-Z][A-Z0-9]{2,}$")   # TERM, STDTC, ORRES - not "E" or "X"


@dataclass
class Issue:
    """One problem with one cell. `code` is stable; `message` is for a human."""
    row_id: str
    column: str
    severity: str
    code: str
    message: str


@dataclass
class ImportedRow:
    """One reviewed annotation, ready for the PDF writer if it passed."""
    row_id: str
    annot_seq: int = 1
    # FIELD | FORM. A FORM row's `row_id` names a page's header band rather than
    # a question, so it is validated against the anchors instead of the fields -
    # everything after that (geometry, placement, drawing) is identical.
    scope: str = FIELD_SCOPE
    form_name: str = ""
    field_text: str = ""
    status: str = ""
    variable: str = ""
    annotation_text: str = ""
    annot_type: str = ""
    # What pre-fill proposed. Kept so a REJECTED row is usable evidence: without
    # the suggestion, "the reviewer said no" says nothing about what to stop
    # suggesting next time.
    suggested_variable: str = ""
    suggested_annotation: str = ""
    match_tier: str = ""
    text_color: tuple[float, float, float] | None = None
    # The box's background. Blank is a decision, not an omission - a study whose
    # house style draws unfilled boxes leaves this empty and gets a transparent
    # box with the same black border as every other one.
    fill_color: tuple[float, float, float] | None = None
    font_name: str = ""
    font_size: float = 0.0
    placement: str = ""
    geometry: dict[str, Any] = dc_field(default_factory=dict)
    reviewer_note: str = ""
    issues: list[Issue] = dc_field(default_factory=list)

    @property
    def key(self) -> tuple[str, int]:
        """Identity in the workbook: which field, and which of its annotations."""
        return (self.row_id, self.annot_seq)

    @property
    def ok(self) -> bool:
        return not any(i.severity == ERROR for i in self.issues)

    @property
    def text_to_draw(self) -> str:
        """What actually gets rendered: the annotation, falling back to the bare
        variable when a reviewer filled only that column."""
        return self.annotation_text or self.variable

    @property
    def ready(self) -> bool:
        """Passed validation *and* a human approved it. Both, or it is not written."""
        return self.ok and self.status == APPROVED


@dataclass
class ImportReport:
    path: str
    rows: list[ImportedRow] = dc_field(default_factory=list)
    issues: list[Issue] = dc_field(default_factory=list)   # workbook-level only

    @property
    def all_issues(self) -> list[Issue]:
        return self.issues + [i for r in self.rows for i in r.issues]

    @property
    def errors(self) -> list[Issue]:
        return [i for i in self.all_issues if i.severity == ERROR]

    @property
    def warnings(self) -> list[Issue]:
        return [i for i in self.all_issues if i.severity == WARNING]

    def approved(self) -> list[ImportedRow]:
        """The rows the PDF writer may act on."""
        return [r for r in self.rows if r.ready]

    def rejected(self) -> list[ImportedRow]:
        """Rows a reviewer turned down, with the suggestion they turned down.

        Negative evidence: the corpus should stop proposing what a human has
        already said no to.
        """
        return [r for r in self.rows
                if r.status == REJECTED and (r.suggested_annotation or r.suggested_variable)]

    def blocked(self) -> list[ImportedRow]:
        return [r for r in self.rows if not r.ok]

    def rows_for(self, row_id: str) -> list[ImportedRow]:
        """Every annotation on one field, in seq order."""
        return sorted((r for r in self.rows if r.row_id == row_id),
                      key=lambda r: r.annot_seq)

    def multi_annotation_fields(self) -> dict[str, int]:
        """field -> how many annotations it carries, for fields carrying several."""
        counts: dict[str, int] = {}
        for r in self.rows:
            counts[r.row_id] = counts.get(r.row_id, 0) + 1
        return {k: v for k, v in sorted(counts.items()) if v > 1}

    def form_rows(self) -> list[ImportedRow]:
        """Rows that annotate a form rather than one of its fields."""
        return [r for r in self.rows if r.scope == FORM_SCOPE]

    def to_dict(self) -> dict[str, Any]:
        return {
            "path": self.path,
            "rows": len(self.rows),
            "fields": len({r.row_id for r in self.rows if r.scope != FORM_SCOPE}),
            "form_rows": len(self.form_rows()),
            "form_rows_approved": len([r for r in self.form_rows() if r.ready]),
            "multi_annotation_fields": len(self.multi_annotation_fields()),
            "approved": len(self.approved()),
            "blocked": len(self.blocked()),
            "errors": len(self.errors),
            "warnings": len(self.warnings),
            "by_code": _counts(i.code for i in self.all_issues),
            "by_status": _counts(r.status for r in self.rows),
        }


def read_staging(path: str | Path, doc: Document | None = None) -> ImportReport:
    """Read and validate a returned staging workbook.

    Pass the re-parsed blank CRF as `doc` for the strongest checks: that every
    row still refers to a field that exists, that none were lost, and that the
    label was not edited out from under its `row_id`.
    """
    path = Path(path)
    report = ImportReport(path=str(path))
    values = load_workbook(path, data_only=True)
    formulas = load_workbook(path, data_only=False)

    if SHEET_WORK not in values.sheetnames:
        report.issues.append(Issue("", SHEET_WORK, ERROR, "MISSING_SHEET",
                                   f"workbook has no {SHEET_WORK!r} sheet"))
        return report

    geometry = _read_geometry(values, report)
    ws, fws = values[SHEET_WORK], formulas[SHEET_WORK]
    names = [c.value for c in ws[1]]

    seen: set[tuple[str, int]] = set()
    used_seqs: dict[str, set[int]] = {}
    # Fields and anchors together: both are things a row_id may name, and both
    # must still exist in the CRF for the row to be placeable. Kept in one dict
    # because every check downstream asks the same question of them - does this
    # id exist, and does the row still describe what it points at.
    fields = {f.id: f for f in doc.iter_fields()} if doc else {}
    if doc:
        fields.update({a.id: a for a in doc.iter_anchors()})
    # Field carries its form's *name*; the domain lives on the Form - and for a
    # blank CRF it is only knowable from history, which the exporter recorded on
    # the Forms sheet. Prefer that, falling back to the document's own.
    domains = {normalize(f.name): f.domain for f in doc.forms if f.domain} if doc else {}
    domains.update(_read_domains(values))
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        raw = {n: c.value for n, c in zip(names, row)}
        raw_f = {n: c.value for n, c in zip(names, fws[i])}
        imported = _row(raw, raw_f, geometry, fields, domains, seen, used_seqs)
        if imported:
            report.rows.append(imported)

    _check_sets(report)
    if doc:
        # Fields only. A field with no rows is a field nobody was asked about,
        # which is an error - but a page carrying no form-level markup is
        # ordinary (a continuation page usually repeats no header), so an anchor
        # with no rows is a legitimate answer, not a lost one. The way to record
        # "this page needs none" and have the corpus learn it is still to REJECT
        # the row rather than delete it.
        wanted = {f.id for f in doc.iter_fields()}
        for missing in sorted(wanted - {rid for rid, _ in seen}):
            report.issues.append(Issue(
                missing, "row_id", ERROR, "MISSING_ROW",
                f"field {missing} ({fields[missing].text!r}) has no row at all - "
                "a field may lose annotations, but not every one of them"))
    return report


def _check_sets(report: ImportReport) -> None:
    """Cross-row checks within one field's set of annotations.

    The one thing a per-row check cannot see: two rows on the same field saying
    the same thing. It reads as thoroughness in the sheet and prints the same
    markup twice on the page, so it is caught here rather than by eye on page 40.
    """
    for row_id in sorted({r.row_id for r in report.rows}):
        rows = report.rows_for(row_id)
        if len(rows) < 2:
            continue
        seen: dict[tuple[str, ...], int] = {}
        for r in rows:
            key = statement_key(r.text_to_draw)
            if not key:
                continue
            if key in seen:
                r.issues.append(Issue(
                    r.row_id, "final_annotation", WARNING, "DUPLICATE_STATEMENT",
                    f"annot_seq {r.annot_seq} says the same thing as seq "
                    f"{seen[key]}; only one of them will be drawn"))
            else:
                seen[key] = r.annot_seq


# --- one row ---------------------------------------------------------------
def _row(raw: dict, formulas: dict, geometry: dict, fields: dict, domains: dict,
         seen: set[tuple[str, int]], used_seqs: dict[str, set[int]]) -> ImportedRow | None:
    row_id = _text(raw.get("row_id"))
    if not row_id:
        return None                       # a blank trailing row is not an error
    r = ImportedRow(row_id=row_id,
                    form_name=_text(raw.get("form_name")),
                    field_text=_text(raw.get("field_text")),
                    suggested_variable=_text(raw.get("suggested_variable")),
                    suggested_annotation=_text(raw.get("suggested_annotation")),
                    match_tier=_text(raw.get("match_tier")),
                    reviewer_note=_text(raw.get("reviewer_note")))
    add = lambda col, sev, code, msg: r.issues.append(Issue(row_id, col, sev, code, msg))

    _check_seq(r, raw, used_seqs, add)
    _check_scope(r, raw, fields, add)
    _check_identity(r, raw, fields, seen, add)
    _check_formulas(r, raw, formulas, add)
    _check_status(r, raw, add)
    _check_decision(r, raw, domains, add)
    _check_formatting(r, raw, add)
    _check_geometry(r, geometry, add)
    return r


def _check_seq(r, raw, used_seqs: dict[str, set[int]], add) -> None:
    """Which annotation of its field this row is.

    Blank is the common case for a row a reviewer added: they copied a row and
    wrote the new statement, which is the part that needed a person. Numbering it
    for them is safe - "the next free one" has exactly one answer - but it is
    still said out loud, because a blank seq is as easily a row they forgot to
    finish as one they meant to leave to us.
    """
    taken = used_seqs.setdefault(r.row_id, set())
    raw_seq = raw.get("annot_seq")
    if raw_seq in (None, ""):
        r.annot_seq = next(n for n in range(1, len(taken) + 2) if n not in taken)
        if taken:
            add("annot_seq", WARNING, "IMPLIED_SEQ",
                f"blank annot_seq on a field that already has {len(taken)} "
                f"annotation(s); numbered {r.annot_seq}")
        taken.add(r.annot_seq)
        return
    try:
        seq = int(float(raw_seq))
    except (TypeError, ValueError):
        add("annot_seq", ERROR, "BAD_ANNOT_SEQ",
            f"{raw_seq!r} is not a whole number")
        return
    if seq < 1:
        add("annot_seq", ERROR, "BAD_ANNOT_SEQ",
            f"annot_seq {seq} is not a positive number")
        return
    r.annot_seq = seq
    taken.add(seq)


def _check_scope(r, raw, fields, add) -> None:
    """Does this row belong to a field or to the form itself?

    The `row_id` is the authority, not the column: it is the thing that resolves
    to a position on the page, and a reviewer who copies a form row onto a field
    row (or the reverse) has changed what the row *is* without meaning to. So a
    disagreement is corrected from the id and reported, rather than believed -
    believing the column would place a domain header on some question's row.
    """
    from .models import FormAnchor
    written = _text(raw.get("scope")).upper()
    target = fields.get(r.row_id)
    actual = (FORM_SCOPE if isinstance(target, FormAnchor)
              else FIELD_SCOPE if target is not None
              else written or FIELD_SCOPE)
    if written and written not in SCOPES:
        add("scope", ERROR, "BAD_SCOPE",
            f"{written!r} is not one of {', '.join(SCOPES)}")
    elif written and target is not None and written != actual:
        add("scope", WARNING, "SCOPE_MISMATCH",
            f"row says {written} but {r.row_id} is a "
            f"{'form header' if actual == FORM_SCOPE else 'field'}; treated as {actual}")
    r.scope = actual


def _check_identity(r, raw, fields, seen, add) -> None:
    """row_id is the only link to a position on the PDF. Guard it hard.

    Rows may be *added* - that is how a field gets a second annotation - so what
    is guarded is not the row count but the key: (row_id, annot_seq) identifies
    one annotation, and two rows claiming it is unresolvable.
    """
    if any(i.code == "BAD_ANNOT_SEQ" for i in r.issues):
        return                            # its key is unknown; do not invent one
    if r.key in seen:
        add("annot_seq", ERROR, "DUPLICATE_ROW_KEY",
            f"{r.row_id} annot_seq {r.annot_seq} appears more than once - give "
            "each annotation of a field its own seq")
        return
    seen.add(r.key)
    if not fields:
        return
    fld = fields.get(r.row_id)
    if fld is None:
        add("row_id", ERROR, "UNKNOWN_ROW_ID",
            f"{r.row_id} is not a field in this CRF - a row added for a second "
            "annotation must keep the row_id it was copied from, and the workbook "
            "must be the one built from this PDF")
    elif normalize(r.field_text) != normalize(fld.text):
        add("field_text", ERROR, "ROW_ALTERED",
            f"field_text {r.field_text!r} no longer matches {fld.text!r}; the row "
            "does not describe the field its row_id points at")
    elif normalize(r.form_name) != normalize(fld.form_name):
        # Expected and useful - but it changes the primary key, so it is recorded
        # rather than accepted silently.
        add("form_name", WARNING, "FORM_RENAMED",
            f"reviewer moved this field from {fld.form_name!r} to {r.form_name!r}; "
            "the mapping will be stored under the new form")


def _check_formulas(r, raw, formulas, add) -> None:
    """An uncalculated formula reads as an empty cell. Say so instead."""
    for col in ("final_variable", "final_annotation", "status"):
        written = formulas.get(col)
        if raw.get(col) in (None, "") and isinstance(written, str) and written.startswith("="):
            add(col, ERROR, "UNEVALUATED_FORMULA",
                f"{col} holds the formula {written!r} with no cached result - open "
                "the workbook in Excel and save it, or paste values")


def _check_status(r, raw, add) -> None:
    r.status = _text(raw.get("status")).upper()
    if not r.status:
        add("status", ERROR, "MISSING_STATUS", "every row needs a status")
    elif r.status not in STATUSES:
        add("status", ERROR, "BAD_STATUS",
            f"{r.status!r} is not one of {', '.join(STATUSES)}")
    elif r.status not in (APPROVED, REJECTED):
        add("status", WARNING, "NOT_REVIEWED",
            f"still {r.status}; only APPROVED rows are written to the PDF")
    if r.status == REJECTED and not r.reviewer_note:
        add("reviewer_note", WARNING, "NO_REJECT_REASON",
            "rejected without a note - the next study cannot learn from this")


def _check_decision(r, raw, domains, add) -> None:
    """The mapping itself: variable, annotation text, and their agreement."""
    r.variable = _text(raw.get("final_variable")).upper()
    r.annotation_text = _text(raw.get("final_annotation")) or r.variable
    r.annot_type = _text(raw.get("final_annot_type")).upper()

    if r.status == APPROVED and not r.annotation_text:
        add("final_annotation", ERROR, "MISSING_DECISION",
            "approved but no annotation to write")
        return
    if not r.annotation_text:
        return

    if r.variable and not _VARIABLE.match(r.variable):
        add("final_variable", ERROR, "BAD_VARIABLE",
            f"{r.variable!r} is not an SDTM variable token")

    # Classify what was actually typed with the same rules the parser uses, so
    # the workbook cannot declare a type its own text contradicts. A FORM row is
    # page-topmost markup by construction, which is the signal that separates
    # "DS=DISPOSITION" (a domain header) from "DS=COMPLETED" (a constant).
    parsed = ann.classify(r.annotation_text, first=r.scope == FORM_SCOPE)
    if not r.annot_type:
        r.annot_type = parsed.annot_type
    elif r.annot_type not in ANNOT_TYPES:
        add("final_annot_type", ERROR, "BAD_ANNOT_TYPE",
            f"{r.annot_type!r} is not a known annotation type")
    elif r.annot_type != parsed.annot_type and parsed.confidence >= ann.CONF_LIKELY:
        add("final_annot_type", WARNING, "TYPE_MISMATCH",
            f"declared {r.annot_type}, but {r.annotation_text!r} reads as "
            f"{parsed.annot_type} ({parsed.evidence[0]})")

    domain = domains.get(normalize(r.form_name), "")
    foreign = _foreign_domain(r.variable, domain)
    if foreign:
        # A warning, never an error - a study may legitimately annotate one form
        # with another domain's variable.
        add("final_variable", WARNING, "DOMAIN_MISMATCH",
            f"{r.variable} looks like a {foreign} variable, but this form is {domain}")


def _foreign_domain(variable: str, domain: str) -> str:
    """Does this variable look like it belongs to a *different* domain?

    "Missing the form's prefix" is the obvious test and it is wrong: DM's own
    variables carry no prefix at all, so BRTHDTC, AGE and RACE would each raise a
    warning on the one form they belong to. Three false alarms in four teaches a
    reviewer to ignore the check, which costs more than the check is worth.

    What actually indicates a mistake is a variable wearing *another* domain's
    code with a real suffix behind it - DSTERM on a Medical History form. The
    suffix length is what keeps AGE ("AG" + "E") and SEX ("SE" + "X") out, since
    both start with letters that happen to be CDISC domain codes.
    """
    if not (variable and domain) or variable.startswith(domain):
        return ""
    prefix, rest = variable[:2], variable[2:]
    if prefix in ann.DOMAINS and _SUFFIX.match(rest):
        return prefix
    return ""


def _check_formatting(r, raw, add) -> None:
    r.text_color = _color(raw, "color_rgb", add)
    r.fill_color = _color(raw, "fill_rgb", add)
    r.font_name = _text(raw.get("font_name"))

    size = raw.get("font_size")
    if size not in (None, ""):
        try:
            r.font_size = float(size)
        except (TypeError, ValueError):
            add("font_size", ERROR, "BAD_FONT_SIZE", f"{size!r} is not a number")
        else:
            if not MIN_FONT_PT <= r.font_size <= MAX_FONT_PT:
                add("font_size", ERROR, "BAD_FONT_SIZE",
                    f"{r.font_size}pt is outside {MIN_FONT_PT}-{MAX_FONT_PT}pt")

    r.placement = _text(raw.get("placement"))
    if r.placement and r.placement not in PLACEMENTS:
        add("placement", ERROR, "BAD_PLACEMENT",
            f"{r.placement!r} is not one of {', '.join(sorted(PLACEMENTS))}")

    if r.status == APPROVED and not r.text_color:
        add("color_rgb", ERROR, "MISSING_COLOR",
            "approved rows need a colour to draw with")


def _color(raw, column: str, add) -> tuple[float, float, float] | None:
    """A #RRGGBB cell as RGB floats. Empty means "no colour", which is allowed."""
    value = _text(raw.get(column))
    if not value:
        return None
    m = _HEX.match(value)
    if not m:
        add(column, ERROR, "BAD_COLOR", f"{value!r} is not a #RRGGBB colour")
        return None
    h = m.group(1)
    return tuple(round(int(h[i:i + 2], 16) / 255, 3) for i in (0, 2, 4))


def _check_geometry(r, geometry, add) -> None:
    """Find this row's geometry, inheriting it from the field where it must.

    A sibling row a reviewer added has no `Geometry` row - that sheet is locked
    and regenerated, and asking them to hand-write page fractions would be worse
    than useless. Every row of one field describes the same box, so the field's
    own geometry is the right answer; only the seq the reviewer chose is new.
    """
    geom = geometry.get(r.key) or _inherited_geometry(geometry, r.row_id)
    if geom is None:
        add("row_id", ERROR, "MISSING_GEOMETRY",
            f"no {SHEET_GEOM} row for {r.row_id} - nowhere to place the annotation")
        return
    for key in GEOM_KEYS:
        value = geom.get(key)
        if not isinstance(value, (int, float)):
            add(key, ERROR, "BAD_GEOMETRY", f"{SHEET_GEOM}.{key} is {value!r}")
            return
        if key.startswith("rel_") and not 0.0 <= value <= 1.0:
            add(key, ERROR, "BAD_GEOMETRY",
                f"{SHEET_GEOM}.{key} = {value} is not a page fraction")
            return
    r.geometry = geom


def _inherited_geometry(geometry: dict, row_id: str) -> dict | None:
    """The lowest-seq geometry recorded for this field."""
    candidates = [(seq, g) for (rid, seq), g in geometry.items() if rid == row_id]
    return min(candidates, key=lambda c: c[0])[1] if candidates else None


def _read_domains(wb) -> dict[str, str]:
    """form -> domain, as resolved when the workbook was written."""
    if SHEET_FORMS not in wb.sheetnames:
        return {}
    ws = wb[SHEET_FORMS]
    names = [c.value for c in ws[1]]
    if "domain" not in names:
        return {}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(names, row))
        if r.get("form_name") and r.get("domain"):
            out[normalize(str(r["form_name"]))] = str(r["domain"]).strip().upper()
    return out


def _read_geometry(wb, report: ImportReport) -> dict[tuple[str, int], dict]:
    """Keyed by (row_id, annot_seq).

    A workbook written before `annot_seq` existed has no such column; its rows
    are read as seq 1, which is what they were.
    """
    if SHEET_GEOM not in wb.sheetnames:
        report.issues.append(Issue("", SHEET_GEOM, ERROR, "MISSING_SHEET",
                                   f"workbook has no {SHEET_GEOM!r} sheet"))
        return {}
    ws = wb[SHEET_GEOM]
    names = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        g = dict(zip(names, row))
        if g.get("row_id"):
            out[(str(g["row_id"]), _seq(g.get("annot_seq")))] = g
    return out


def _seq(value, default: int = 1) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


# --- helpers ---------------------------------------------------------------
def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _counts(values) -> dict[str, int]:
    out: dict[str, int] = {}
    for v in values:
        out[v] = out.get(v, 0) + 1
    return dict(sorted(out.items()))


def write_review_copy(report: ImportReport, source: str | Path,
                      path: str | Path) -> Path:
    """Write the workbook back with an `import_issues` column filled in.

    Closes the loop in the tool the reviewer is already in: every blocked row
    says what to fix, in the row that needs fixing, instead of in a log they
    would have to cross-reference by hand.
    """
    from openpyxl.styles import Font, PatternFill
    wb = load_workbook(source)
    ws = wb[SHEET_WORK]
    names = [c.value for c in ws[1]]
    col = names.index("import_issues") + 1 if "import_issues" in names else len(names) + 1
    ws.cell(row=1, column=col, value="import_issues").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=1, column=col).fill = PatternFill("solid", fgColor="263238")
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 60

    # Paired by position, not by key. `report.rows` was built by walking these
    # same rows in this same order, so the n-th non-blank row is the n-th report
    # row - and that stays true for a row whose seq was blank and got numbered on
    # the way in, which a key lookup could not find its way back to. The row_id
    # is checked anyway, so a mismatch degrades to "no verdict" rather than to
    # somebody else's verdict.
    id_col = names.index("row_id") + 1
    pending = list(report.rows)
    for i in range(2, ws.max_row + 1):
        row_id = _text(ws.cell(row=i, column=id_col).value)
        if not row_id:
            continue
        r = pending.pop(0) if pending else None
        if not r or r.row_id != row_id or not r.issues:
            continue
        ws.cell(row=i, column=col, value="; ".join(
            f"[{x.severity}] {x.column}: {x.message}" for x in r.issues))
        if not r.ok:
            ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="FFEBEE")
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
