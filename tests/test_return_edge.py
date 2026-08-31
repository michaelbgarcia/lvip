"""The return edge: approved work flowing back into the knowledge base."""
import json

import pytest
from openpyxl import load_workbook

from acrf_parser import build_kb
from acrf_parser import staging as st
from acrf_parser.importer import read_staging
from acrf_parser.kb import KnowledgeBase, ingest_approved
from acrf_parser.models import HUMAN_APPROVED
from acrf_parser.prefill import PrefillIndex, prefill_document, summarize_prefill
from acrf_parser.style import derive_house_style_from_kb


@pytest.fixture
def corpus_db(corpus, tmp_path):
    path = tmp_path / "corpus.sqlite"
    for doc in corpus:
        build_kb(doc, path)
    return path


@pytest.fixture
def reviewed(blank_doc, corpus_db, tmp_path):
    """A staging sheet as it comes back: gaps filled, a form renamed, one rejection."""
    with KnowledgeBase(corpus_db) as kb:
        st.write_staging(blank_doc, tmp_path / "s.xlsx",
                         index=PrefillIndex.from_kb(kb),
                         house=derive_house_style_from_kb(kb))
    wb = load_workbook(tmp_path / "s.xlsx")
    ws = wb[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    col = lambda n: names.index(n) + 1

    filled = {"p4f0": ("DSTERM", "Disposition"), "p5f0": ("IETESTCD", None),
              "p5f1": ("IETESTCD", None), "p5f2": ("IETESTCD", None),
              "p5f3": ("IETESTCD", None)}
    for i in range(2, ws.max_row + 1):
        rid = ws.cell(row=i, column=col("row_id")).value
        ws.cell(row=i, column=col("status")).value = "APPROVED"
        if rid in filled:
            var, form = filled[rid]
            ws.cell(row=i, column=col("final_variable")).value = var
            ws.cell(row=i, column=col("final_annotation")).value = var
            if form:
                ws.cell(row=i, column=col("form_name")).value = form
        if rid == "p2f2":
            ws.cell(row=i, column=col("status")).value = "REJECTED"
            ws.cell(row=i, column=col("reviewer_note")).value = "wrong on this study"
    out = tmp_path / "reviewed.xlsx"
    wb.save(out)
    return read_staging(out, blank_doc)


def test_approved_work_raises_the_next_run(blank_doc, corpus_db, reviewed):
    """The point of the whole edge: the expensive rows are paid for once.

    Five rows reached the agent before; two after - and both survivors are
    correct rather than missed:

    * "Ongoing" was *rejected* by the reviewer, so it is deliberately no longer
      suggested (see the negative-evidence tests below).
    * The Disposition question was filed under the form name the reviewer
      corrected it to, while this blank CRF still parses that page as Medical
      History. The knowledge is stored right; the key simply does not match
      until a document identifies the form correctly. That is the limitation of
      a per-document correction, and it is honest rather than hidden.
    """
    with KnowledgeBase(corpus_db) as kb:
        before = summarize_prefill(prefill_document(blank_doc, PrefillIndex.from_kb(kb)))
    assert before["by_status"]["NEEDS_MAPPING"] == 5

    ingest_approved(reviewed, blank_doc, corpus_db, source="STUDY-NEW")
    with KnowledgeBase(corpus_db) as kb:
        after = summarize_prefill(prefill_document(blank_doc, PrefillIndex.from_kb(kb)))
    assert after["by_status"]["NEEDS_MAPPING"] == 2
    assert after["auto_fill_rate"] > before["auto_fill_rate"]


def test_ingest_does_not_touch_the_callers_document(blank_doc, corpus_db, reviewed):
    """Ingest renames fields and replaces annotation lists; doing that to the
    caller's parse would silently rewrite a document they are still holding."""
    before_names = [(f.id, f.form_name) for f in blank_doc.iter_fields()]
    before_annots = [len(p.annotations) for p in blank_doc.pages]
    ingest_approved(reviewed, blank_doc, corpus_db, source="STUDY-NEW")
    assert [(f.id, f.form_name) for f in blank_doc.iter_fields()] == before_names
    assert [len(p.annotations) for p in blank_doc.pages] == before_annots


def test_the_sheet_form_name_is_believed_over_the_parse(blank_doc, corpus_db, reviewed):
    """A blank CRF has no domain headers, so its Disposition page reads as
    "Medical History". Re-parsing the written PDF would relearn that mistake;
    the reviewer's correction is what gets stored."""
    ingest_approved(reviewed, blank_doc, corpus_db, source="STUDY-NEW")
    question = "Please record protocol version on which subject is currently enrolled"
    with KnowledgeBase(corpus_db) as kb:
        rows = kb.lookup("Disposition", question)
        # Both occurrences are kept - the original study annotated this question
        # SUPPDS.QVAL - but the reviewer's decision ranks first, which is what
        # `variable_for` returns and what pre-fill will offer.
        assert [r["variable"] for r in rows] == ["DSTERM", "QVAL"]
        assert rows[0]["trust"] == HUMAN_APPROVED
        assert kb.variable_for("Disposition", question) == "DSTERM"
        # And nothing was filed under the form the blank CRF guessed.
        assert kb.variable_for("Medical History", question) is None


def test_a_rename_is_recorded_not_silently_accepted(reviewed):
    renames = [i for i in reviewed.warnings if i.code == "FORM_RENAMED"]
    assert len(renames) == 1
    assert "Medical History" in renames[0].message and "Disposition" in renames[0].message


def test_form_name_is_editable_in_the_workbook():
    assert "form_name" in st.EDITABLE


def test_approved_links_are_marked_as_human(blank_doc, corpus_db, reviewed):
    ingest_approved(reviewed, blank_doc, corpus_db, source="STUDY-NEW")
    with KnowledgeBase(corpus_db) as kb:
        trusts = {r["trust"] for r in kb.con.execute(
            "SELECT trust FROM field_annotations WHERE file_name = 'STUDY-NEW'")}
    assert trusts == {HUMAN_APPROVED}


def test_a_human_decision_outranks_a_geometric_one(blank_doc, corpus_db, reviewed):
    """A reviewer's answer must not lose a tie to a lucky geometric match."""
    ingest_approved(reviewed, blank_doc, corpus_db, source="STUDY-NEW")
    with KnowledgeBase(corpus_db) as kb:
        index = PrefillIndex.from_kb(kb)
    result = {r.field_id: r for r in prefill_document(blank_doc, index)}
    best = result["p1f0"].best
    assert best.trust == HUMAN_APPROVED and best.confidence == 1.0
    assert "approved by a reviewer" in best.evidence[0]


# --- negative evidence -----------------------------------------------------
def test_rejections_are_captured_with_what_was_rejected(reviewed):
    """"The reviewer said no" is useless without what they said no *to*."""
    rejected = reviewed.rejected()
    assert [r.row_id for r in rejected] == ["p2f2"]
    assert rejected[0].suggested_annotation == "MHENRF=ONGOING"
    assert rejected[0].match_tier == "EXACT_KEY"


def test_a_rejected_suggestion_is_not_offered_again(blank_doc, corpus_db, reviewed):
    """Re-proposing what a human already refused burns their trust in every
    other row on the sheet."""
    ingest_approved(reviewed, blank_doc, corpus_db, source="STUDY-NEW")
    with KnowledgeBase(corpus_db) as kb:
        assert [r["annotation_text"] for r in kb.con.execute(
            "SELECT annotation_text FROM rejected_suggestions")] == ["MHENRF=ONGOING"]
        index = PrefillIndex.from_kb(kb)

    result = {r.field_id: r for r in prefill_document(blank_doc, index)}
    ongoing = result["p2f2"].best
    assert ongoing.confidence == 0.0
    assert "a reviewer rejected this suggestion previously" in ongoing.evidence
    assert result["p2f2"].status == "NEEDS_MAPPING"


def test_rejected_rows_are_never_written_to_the_pdf(reviewed):
    assert all(r.row_id != "p2f2" for r in reviewed.approved())


# --- a set, all the way round ----------------------------------------------
@pytest.fixture
def reviewed_with_a_set(blank_doc, corpus_db, tmp_path):
    """The same sheet, with a reviewer adding three more annotations to one field.

    "Date of birth" arrives pre-filled with BRTHDTC alone; the reviewer says it
    also carries RFICDTC and DSSTDTC, the way a real consent date does.
    """
    with KnowledgeBase(corpus_db) as kb:
        st.write_staging(blank_doc, tmp_path / "set.xlsx",
                         index=PrefillIndex.from_kb(kb),
                         house=derive_house_style_from_kb(kb))
    wb = load_workbook(tmp_path / "set.xlsx")
    ws = wb[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    col = lambda n: names.index(n) + 1
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=col("status")).value == "AUTO":
            ws.cell(row=i, column=col("status")).value = "APPROVED"

    source = next(i for i in range(2, ws.max_row + 1)
                  if ws.cell(row=i, column=col("row_id")).value == "p1f0")
    for seq, text in enumerate(["RFICDTC", "DSSTDTC"], start=2):
        ws.append([ws.cell(row=source, column=j).value for j in range(1, len(names) + 1)])
        for name, value in (("annot_seq", seq), ("final_variable", text),
                            ("final_annotation", text), ("status", "APPROVED")):
            ws.cell(row=ws.max_row, column=col(name)).value = value
    out = tmp_path / "reviewed_set.xlsx"
    wb.save(out)
    return read_staging(out, blank_doc)


def test_an_added_set_survives_the_whole_round_trip(blank_doc, corpus_db,
                                                    reviewed_with_a_set, tmp_path):
    """Reviewer adds two annotations -> drawn on the PDF -> learned -> proposed
    back as a set on the next study. Every stage has to carry all three or the
    loop quietly loses the reviewer's work."""
    from acrf_parser.writer import write_annotations

    report = write_annotations(blank_doc.path, reviewed_with_a_set.rows,
                               tmp_path / "set.pdf")
    drawn = {p.text for p in report.placements if p.row_id == "p1f0"}
    assert drawn == {"BRTHDTC", "RFICDTC", "DSSTDTC"}

    ingest_approved(reviewed_with_a_set, blank_doc, corpus_db, source="STUDY-SET")
    with KnowledgeBase(corpus_db) as kb:
        index = PrefillIndex.from_kb(kb)
    again = {r.field_id: r for r in prefill_document(blank_doc, index)}["p1f0"]
    # In the reviewer's own order, not alphabetically: the set is ranked by where
    # it was drawn, and it was drawn left to right in annot_seq order.
    assert [c.annotation_text for c in again.annotations] == [
        "BRTHDTC", "RFICDTC", "DSSTDTC"]
    assert all(again.status_of(c) == "AUTO" for c in again.annotations)


def test_the_learned_set_is_not_stacked_at_one_point(blank_doc, corpus_db,
                                                     reviewed_with_a_set):
    """Geometry is stored where the writer would draw it. Filing three
    annotations at the same x would teach the next study an offset nobody used."""
    ingest_approved(reviewed_with_a_set, blank_doc, corpus_db, source="STUDY-SET")
    with KnowledgeBase(corpus_db) as kb:
        boxes = [json.loads(r[0]) for r in kb.con.execute(
            "SELECT a.bbox FROM annotations a JOIN documents d ON a.document_id = d.id"
            " WHERE d.file_name = 'STUDY-SET' AND a.text IN"
            " ('BRTHDTC', 'RFICDTC', 'DSSTDTC')")]
    xs = sorted(b[0] for b in boxes)
    assert len(xs) == 3 and len(set(xs)) == 3
    assert all(b < a for a, b in zip(xs[1:], xs)), "laid out left to right"


def test_house_style_learns_from_the_approved_geometry(blank_doc, corpus_db, reviewed):
    """The annotations we store sit where the writer would draw them, so the
    next study's placement convention reflects what was actually used."""
    ingest_approved(reviewed, blank_doc, corpus_db, source="STUDY-NEW")
    with KnowledgeBase(corpus_db) as kb:
        house = derive_house_style_from_kb(kb)
    assert house.default.placement == "right_of_field"
    assert house.for_type("VARIABLE").samples > 10
