"""Phase 9c tests: reading the staging workbook back, per row."""
import pytest
from openpyxl import load_workbook

from acrf_parser import importer as imp
from acrf_parser import staging as st
from acrf_parser.importer import WARNING, read_staging, write_review_copy


@pytest.fixture
def returned(blank_doc, index, house, tmp_path):
    """A workbook that has been out to Excel and edited - well and badly."""
    path = st.write_staging(blank_doc, tmp_path / "out.xlsx", index=index, house=house)
    wb = load_workbook(path)
    ws = wb[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    col = lambda n: names.index(n) + 1

    def edit(row, **cells):
        # Assign through .value: cell(..., value=None) is a no-op in openpyxl,
        # so the obvious spelling silently fails to clear a cell.
        for name, value in cells.items():
            ws.cell(row=row, column=col(name)).value = value

    for i in range(2, ws.max_row + 1):                 # reviewer approves the AUTO rows
        if ws.cell(row=i, column=col("status")).value == "AUTO":
            edit(i, status="APPROVED")
    def find(row_id, seq=1):
        """Sheet row holding one annotation. Tests must not assume an order:
        each page now leads with its form-level rows, so row 2 is a form header."""
        for i in range(2, ws.max_row + 1):
            if (ws.cell(row=i, column=col("row_id")).value == row_id
                    and int(ws.cell(row=i, column=col("annot_seq")).value or 1) == seq):
                return i
        raise AssertionError(f"no row for {row_id}#{seq}")

    ws.find = find
    out = tmp_path / "returned.xlsx"
    wb.save(out)
    return out, wb, ws, edit


def _read(path, doc):
    return read_staging(path, doc)


def _geom_row(wb, row_id, seq=1):
    ws = wb[st.SHEET_GEOM]
    names = [c.value for c in ws[1]]
    rid, sq = names.index("row_id") + 1, names.index("annot_seq") + 1
    for i in range(2, ws.max_row + 1):
        if (ws.cell(row=i, column=rid).value == row_id
                and int(ws.cell(row=i, column=sq).value or 1) == seq):
            return i
    raise AssertionError(f"no geometry row for {row_id}#{seq}")


# The blank CRF has 14 fields, and one form-header row per page that belongs to
# a form. Counts here are fields + form rows, never fields alone.
FIELD_ROWS, FORM_ROWS = 14, 5
ALL_ROWS = FIELD_ROWS + FORM_ROWS


def test_a_clean_return_imports_whole(returned, blank_doc):
    path, wb, ws, _ = returned
    report = _read(path, blank_doc)
    assert len(report.rows) == ALL_ROWS
    # 9 field rows plus every page's form header, which history also answered
    assert len(report.approved()) == 9 + FORM_ROWS
    assert not report.errors


def test_rows_are_matched_by_row_id_not_position(returned, blank_doc, tmp_path):
    """Sorting a column in Excel must not shift annotations onto other fields."""
    path, wb, ws, _ = returned
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for i, values in enumerate(reversed(rows), start=2):     # reverse every row
        for j, v in enumerate(values, start=1):
            ws.cell(row=i, column=j, value=v)
    shuffled = tmp_path / "shuffled.xlsx"
    wb.save(shuffled)
    report = _read(shuffled, blank_doc)
    assert not report.errors
    by_id = {r.row_id: r for r in report.rows}
    assert by_id["p1f0"].annotation_text == "BRTHDTC"
    assert by_id["p2f1"].annotation_text == "MHSTDTC"


def test_one_bad_row_does_not_block_the_others(returned, blank_doc, tmp_path):
    """The whole point of per-row validation."""
    path, wb, ws, edit = returned
    edit(ws.find("p1f0"), color_rgb="crimson")   # row p1f0 only
    broken = tmp_path / "broken.xlsx"
    wb.save(broken)
    report = _read(broken, blank_doc)
    blocked = report.blocked()
    assert [r.row_id for r in blocked] == ["p1f0"]
    assert len(report.approved()) == 8 + FORM_ROWS   # everything else imports


# --- identity --------------------------------------------------------------
def test_deleted_row_is_caught(returned, blank_doc, tmp_path):
    path, wb, ws, _ = returned
    ws.delete_rows(ws.find("p1f0"))
    out = tmp_path / "deleted.xlsx"
    wb.save(out)
    report = _read(out, blank_doc)
    assert [i.code for i in report.issues] == ["MISSING_ROW"]
    assert "p1f0" in report.issues[0].message


def test_unknown_row_id_is_caught(returned, blank_doc, tmp_path):
    path, wb, ws, edit = returned
    edit(ws.find("p1f0"), row_id="p9f9")
    out = tmp_path / "unknown.xlsx"
    wb.save(out)
    codes = {i.code for i in _read(out, blank_doc).errors}
    assert "UNKNOWN_ROW_ID" in codes and "MISSING_ROW" in codes


def test_duplicate_row_key_is_caught(returned, blank_doc, tmp_path):
    """A repeated row_id is fine; a repeated (row_id, annot_seq) is not."""
    path, wb, ws, edit = returned
    edit(ws.find("p1f1"), row_id="p1f0")        # both rows now claim p1f0 seq 1
    out = tmp_path / "dupe.xlsx"
    wb.save(out)
    assert "DUPLICATE_ROW_KEY" in {i.code for i in _read(out, blank_doc).errors}


def test_editing_the_label_out_from_under_its_row_id_is_caught(returned, blank_doc, tmp_path):
    path, wb, ws, edit = returned
    edit(ws.find("p1f0"), field_text="Age of subject")
    out = tmp_path / "altered.xlsx"
    wb.save(out)
    row = next(r for r in _read(out, blank_doc).rows if r.row_id == "p1f0")
    assert not row.ok
    assert [i.code for i in row.issues] == ["ROW_ALTERED"]


# --- decisions -------------------------------------------------------------
def test_approved_without_an_annotation_is_blocked(returned, blank_doc, tmp_path):
    path, wb, ws, edit = returned
    edit(ws.find("p1f0"), final_variable=None, final_annotation=None)
    out = tmp_path / "empty.xlsx"
    wb.save(out)
    row = next(r for r in _read(out, blank_doc).rows if r.row_id == "p1f0")
    assert "MISSING_DECISION" in {i.code for i in row.issues} and not row.ready


def test_unreviewed_rows_warn_and_are_not_written(returned, blank_doc):
    path, wb, ws, _ = returned
    report = _read(path, blank_doc)
    pending = [r for r in report.rows if r.status == "NEEDS_MAPPING"]
    assert len(pending) == 5
    assert all(not r.ready for r in pending)
    assert all("NOT_REVIEWED" in {i.code for i in r.issues} for r in pending)


def test_malformed_variable_is_an_error(returned, blank_doc, tmp_path):
    path, wb, ws, edit = returned
    edit(ws.find("p1f0"), final_variable="brthdtc!")
    out = tmp_path / "badvar.xlsx"
    wb.save(out)
    row = next(r for r in _read(out, blank_doc).rows if r.row_id == "p1f0")
    assert "BAD_VARIABLE" in {i.code for i in row.issues} and not row.ok


def test_declared_type_is_checked_against_the_text(returned, blank_doc, tmp_path):
    """The workbook cannot claim a type its own annotation text contradicts."""
    path, wb, ws, edit = returned
    edit(ws.find("p1f0"), final_annot_type="NOT_SUBMITTED")   # but the text is "BRTHDTC"
    out = tmp_path / "type.xlsx"
    wb.save(out)
    row = next(r for r in _read(out, blank_doc).rows if r.row_id == "p1f0")
    issue = next(i for i in row.issues if i.code == "TYPE_MISMATCH")
    assert issue.severity == WARNING and row.ready       # flagged, still imported
    assert "reads as VARIABLE" in issue.message


@pytest.mark.parametrize("variable,domain,foreign", [
    ("BRTHDTC", "DM", ""), ("AGE", "DM", ""), ("SEX", "DM", ""), ("RACE", "DM", ""),
    ("USUBJID", "MH", ""), ("MHTERM", "MH", ""), ("VSORRES", "VS", ""),
    ("DSTERM", "MH", "DS"), ("AESTDTC", "MH", "AE"), ("CMTRT", "MH", "CM"),
])
def test_only_genuinely_foreign_variables_warn(variable, domain, foreign):
    """DM's own variables carry no prefix; warning on them would be noise that
    teaches reviewers to ignore the check."""
    assert imp._foreign_domain(variable, domain) == foreign


# --- formatting ------------------------------------------------------------
@pytest.mark.parametrize("column,value,code", [
    ("color_rgb", "crimson", "BAD_COLOR"),
    ("color_rgb", "#GG0000", "BAD_COLOR"),
    ("font_size", 96, "BAD_FONT_SIZE"),
    ("font_size", "big", "BAD_FONT_SIZE"),
    ("placement", "somewhere", "BAD_PLACEMENT"),
    ("status", "MAYBE", "BAD_STATUS"),
])
def test_formatting_errors(returned, blank_doc, tmp_path, column, value, code):
    path, wb, ws, edit = returned
    edit(ws.find("p1f0"), **{column: value})
    out = tmp_path / f"{code}.xlsx"
    wb.save(out)
    row = next(r for r in _read(out, blank_doc).rows if r.row_id == "p1f0")
    assert code in {i.code for i in row.issues} and not row.ok


def test_colour_is_resolved_for_the_writer(returned, blank_doc):
    path, _, _, _ = returned
    row = next(r for r in _read(path, blank_doc).rows if r.row_id == "p1f0")
    # #D91A1A back to floats. The 8-bit round trip quantises 0.85/0.1 slightly;
    # that drift is invisible on the page and stable across re-exports.
    assert row.text_color == (0.851, 0.102, 0.102)
    assert row.font_name == "Helv" and row.font_size == 8.0


def test_geometry_reaches_the_writer(returned, blank_doc):
    path, _, _, _ = returned
    row = next(r for r in _read(path, blank_doc).rows if r.row_id == "p1f0")
    assert row.geometry["page_width"] > 1
    assert 0 < row.geometry["rel_x_pct"] < 1


def test_missing_geometry_row_is_an_error(returned, blank_doc, tmp_path):
    path, wb, ws, _ = returned
    wb[st.SHEET_GEOM].delete_rows(_geom_row(wb, "p1f0"))
    out = tmp_path / "nogeom.xlsx"
    wb.save(out)
    row = next(r for r in _read(out, blank_doc).rows if r.row_id == "p1f0")
    assert "MISSING_GEOMETRY" in {i.code for i in row.issues}


def test_uncalculated_formula_is_not_read_as_a_blank(returned, blank_doc, tmp_path):
    """If the agent writes a formula and nobody opens the file in Excel, the
    cached value is empty - which would otherwise import as "no decision"."""
    path, wb, ws, edit = returned
    edit(ws.find("p1f0"), final_annotation="=CONCAT(ZZ2,\"\")")
    out = tmp_path / "formula.xlsx"
    wb.save(out)
    row = next(r for r in _read(out, blank_doc).rows if r.row_id == "p1f0")
    assert "UNEVALUATED_FORMULA" in {i.code for i in row.issues} and not row.ok


# --- several annotations on one field --------------------------------------
def _add_sibling(ws, source_row, **cells):
    """What a reviewer does in Excel: copy a row, paste it under, edit it."""
    names = [c.value for c in ws[1]]
    values = [ws.cell(row=source_row, column=j).value for j in range(1, len(names) + 1)]
    ws.append(values)
    for name, value in cells.items():
        ws.cell(row=ws.max_row, column=names.index(name) + 1).value = value
    return ws.max_row


def test_an_added_sibling_row_is_accepted(returned, blank_doc, tmp_path):
    """The whole point: a field that needs a second annotation gets a second row."""
    path, wb, ws, _ = returned
    _add_sibling(ws, ws.find("p1f0"), annot_seq=2, final_variable="RFICDTC",
                 final_annotation="RFICDTC", status="APPROVED")
    out = tmp_path / "sibling.xlsx"
    wb.save(out)
    report = _read(out, blank_doc)
    assert not report.errors
    rows = report.rows_for("p1f0")
    assert [r.annot_seq for r in rows] == [1, 2]
    assert [r.annotation_text for r in rows] == ["BRTHDTC", "RFICDTC"]
    assert all(r.ready for r in rows)


def test_a_sibling_inherits_its_field_s_geometry(returned, blank_doc, tmp_path):
    """The Geometry sheet is locked and regenerated - a reviewer cannot add to it,
    and does not need to: every row of a field describes the same box."""
    path, wb, ws, _ = returned
    _add_sibling(ws, ws.find("p1f0"), annot_seq=2, final_annotation="RFICDTC", status="APPROVED")
    out = tmp_path / "inherit.xlsx"
    wb.save(out)
    first, second = _read(out, blank_doc).rows_for("p1f0")
    assert second.geometry and second.geometry["rel_x_pct"] == first.geometry["rel_x_pct"]


def test_a_blank_seq_is_numbered_and_said_out_loud(returned, blank_doc, tmp_path):
    path, wb, ws, _ = returned
    _add_sibling(ws, ws.find("p1f0"), annot_seq=None, final_annotation="RFICDTC", status="APPROVED")
    out = tmp_path / "blankseq.xlsx"
    wb.save(out)
    report = _read(out, blank_doc)
    second = report.rows_for("p1f0")[1]
    assert second.annot_seq == 2 and second.ready
    assert "IMPLIED_SEQ" in {i.code for i in second.issues}


def test_two_rows_saying_the_same_thing_are_flagged(returned, blank_doc, tmp_path):
    """Only one will ever be drawn; better said here than found by eye on page 40."""
    path, wb, ws, _ = returned
    _add_sibling(ws, ws.find("p1f0"), annot_seq=2)                 # an unedited copy: still BRTHDTC
    out = tmp_path / "samestatement.xlsx"
    wb.save(out)
    report = _read(out, blank_doc)
    second = report.rows_for("p1f0")[1]
    issue = next(i for i in second.issues if i.code == "DUPLICATE_STATEMENT")
    assert issue.severity == WARNING and "seq 1" in issue.message


def test_a_field_may_lose_an_annotation_but_not_all_of_them(returned, blank_doc, tmp_path):
    path, wb, ws, _ = returned
    _add_sibling(ws, ws.find("p1f0"), annot_seq=2, final_annotation="RFICDTC", status="APPROVED")
    ws.delete_rows(ws.find("p1f0"))                   # seq 1 gone, seq 2 remains
    out = tmp_path / "onegone.xlsx"
    wb.save(out)
    report = _read(out, blank_doc)
    assert not any(i.code == "MISSING_ROW" and i.row_id == "p1f0"
                   for i in report.issues)
    assert [r.annot_seq for r in report.rows_for("p1f0")] == [2]


def test_the_review_copy_lands_on_the_right_sibling(returned, blank_doc, tmp_path):
    """Two rows share a row_id; a verdict written onto the wrong one is worse
    than none at all."""
    path, wb, ws, _ = returned
    added = _add_sibling(ws, ws.find("p1f0"), annot_seq=2, final_annotation="RFICDTC",
                         status="APPROVED", color_rgb="crimson")
    out = tmp_path / "twoverdicts.xlsx"
    wb.save(out)
    report = _read(out, blank_doc)
    copy = write_review_copy(report, out, tmp_path / "verdicts.xlsx")

    ws2 = load_workbook(copy)[st.SHEET_WORK]
    names = [c.value for c in ws2[1]]
    col = names.index("import_issues") + 1
    assert not ws2.cell(row=ws.find("p1f0"), column=col).value   # seq 1 was fine
    assert "crimson" in ws2.cell(row=added, column=col).value


def test_the_report_counts_annotations_and_fields_separately(returned, blank_doc, tmp_path):
    path, wb, ws, _ = returned
    _add_sibling(ws, ws.find("p1f0"), annot_seq=2, final_annotation="RFICDTC", status="APPROVED")
    out = tmp_path / "counts.xlsx"
    wb.save(out)
    d = _read(out, blank_doc).to_dict()
    assert d["rows"] == ALL_ROWS + 1 and d["fields"] == FIELD_ROWS
    assert d["form_rows"] == FORM_ROWS
    assert d["multi_annotation_fields"] == 1 and d["approved"] == 10 + FORM_ROWS


def test_a_workbook_without_the_seq_column_still_imports(returned, blank_doc, tmp_path):
    """A sheet exported before annot_seq existed is a one-annotation-per-field
    sheet, which is exactly what it now reads as."""
    path, wb, ws, _ = returned
    for sheet in (st.SHEET_WORK, st.SHEET_GEOM):
        names = [c.value for c in wb[sheet][1]]
        wb[sheet].delete_cols(names.index("annot_seq") + 1)
    out = tmp_path / "legacy.xlsx"
    wb.save(out)
    report = _read(out, blank_doc)
    assert not report.errors
    assert len(report.rows) == ALL_ROWS and {r.annot_seq for r in report.rows} == {1}
    assert all(r.geometry for r in report.rows)


def test_several_added_rows_with_no_seq_are_numbered_in_order(returned, blank_doc,
                                                              tmp_path):
    path, wb, ws, _ = returned
    for text in ("RFICDTC", "DSSTDTC"):
        _add_sibling(ws, ws.find("p1f0"), annot_seq=None, final_variable=text,
                     final_annotation=text, status="APPROVED")
    out = tmp_path / "unnumbered.xlsx"
    wb.save(out)
    rows = _read(out, blank_doc).rows_for("p1f0")
    assert [r.annot_seq for r in rows] == [1, 2, 3]
    assert [r.annotation_text for r in rows] == ["BRTHDTC", "RFICDTC", "DSSTDTC"]


# --- reporting -------------------------------------------------------------
def test_report_summary(returned, blank_doc):
    path, _, _, _ = returned
    d = _read(path, blank_doc).to_dict()
    assert d["rows"] == ALL_ROWS and d["blocked"] == 0
    assert d["approved"] == 9 + FORM_ROWS and d["form_rows_approved"] == FORM_ROWS
    assert d["by_status"] == {"APPROVED": 9 + FORM_ROWS, "NEEDS_MAPPING": 5}


def test_review_copy_puts_the_problem_in_the_row(returned, blank_doc, tmp_path):
    """The reviewer is already in Excel; tell them there, not in a log."""
    path, wb, ws, edit = returned
    edit(ws.find("p1f0"), color_rgb="crimson")
    broken = tmp_path / "broken.xlsx"
    wb.save(broken)
    report = _read(broken, blank_doc)
    out = write_review_copy(report, broken, tmp_path / "reviewed.xlsx")

    ws2 = load_workbook(out)[st.SHEET_WORK]
    names = [c.value for c in ws2[1]]
    assert "import_issues" in names
    text = ws2.cell(row=ws.find("p1f0"), column=names.index("import_issues") + 1).value
    assert "[ERROR]" in text and "crimson" in text


def test_importing_without_the_source_pdf_still_validates(returned):
    """`doc` is optional: content checks work standalone, identity checks need it."""
    path, _, _, _ = returned
    report = read_staging(path)
    assert len(report.rows) == ALL_ROWS and len(report.approved()) == 9 + FORM_ROWS
    assert not any(i.code == "MISSING_ROW" for i in report.issues)
