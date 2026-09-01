"""Run the whole pipeline over the MSG pair, so it can be scored.

The loop the tool is for, with the annotated MSG CRF standing in for history:

    annotated aCRF -> corpus (pre-fill index + house style)
    blank CRF      -> staging workbook
    workbook       -> approved rows            (every filled row signed off)
    approved rows  -> annotations on the blank CRF

The approval step is mechanical here, and that is the point: a scored run must
not depend on a person, so everything history could fill is approved as-is.
What that measures is exactly what the pipeline decided on its own - a row
history could not reach stays NEEDS_MAPPING and is never drawn, and shows up in
the score as a miss rather than being quietly hand-finished.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from acrf_parser import parse_pdf, staging as st
from acrf_parser.importer import ImportReport, read_staging
from acrf_parser.models import Document
from acrf_parser.prefill import PrefillIndex
from acrf_parser.style import HouseStyle, derive_house_style
from acrf_parser.writer import WriteReport, write_annotations

from .msg import ANNOTATED, BLANK

APPROVABLE = ("AUTO", "NEEDS_REVIEW")   # everything pre-fill actually filled


@dataclass
class Run:
    """One end-to-end pass, with every intermediate kept for inspection."""
    annotated: Document
    blank: Document
    index: PrefillIndex
    house: HouseStyle
    workbook: Path
    imported: ImportReport
    written: WriteReport
    out_pdf: Path

    @property
    def placements(self):
        return self.written.placements

    def reread(self) -> list:
        """The output PDF's annotations, read back by the parser.

        What the score is actually computed over. The write report says where
        the writer *meant* to put each box; this says what a reader finds on the
        page, and it carries the rendered colour, font and size that the report
        does not - so the same pass that measures placement measures whether the
        style survived being drawn.
        """
        return list(parse_pdf(self.out_pdf).iter_annotations())


def corpus(path: str | Path = ANNOTATED) -> tuple[Document, PrefillIndex, HouseStyle]:
    """Ingest the annotated MSG CRF as this study's history, in memory."""
    doc = parse_pdf(path)
    return doc, PrefillIndex.from_documents([doc]), derive_house_style([doc])


def kb_corpus(tmp: Path, path: str | Path = ANNOTATED
              ) -> tuple[Document, PrefillIndex, HouseStyle]:
    """The same history, round-tripped through SQLite.

    The path the CLI actually takes: `--db` on the way in, `--corpus` on the way
    out, with the PDFs never opened again. It has to reach the same answers as
    the in-memory path or a corpus built on Monday would annotate differently
    from one built in a single run - and the two are separate implementations of
    every lookup, so nothing but a test keeps them honest.
    """
    from acrf_parser.kb import KnowledgeBase, build_kb
    from acrf_parser.style import derive_house_style_from_kb

    doc = parse_pdf(path)
    db = Path(tmp) / "corpus.sqlite"
    build_kb(doc, db)
    with KnowledgeBase(db) as kb:
        return doc, PrefillIndex.from_kb(kb), derive_house_style_from_kb(kb)


def run(tmp: Path, annotated: str | Path = ANNOTATED,
        blank: str | Path = BLANK, via_kb: bool = False) -> Run:
    """Stage the blank CRF against the annotated one, approve, draw."""
    tmp = Path(tmp)
    tmp.mkdir(parents=True, exist_ok=True)
    ann_doc, index, house = (kb_corpus(tmp, annotated) if via_kb
                             else corpus(annotated))
    blank_doc = parse_pdf(blank)

    book = st.write_staging(blank_doc, tmp / "staging.xlsx", index=index, house=house)
    approved_book = approve(book, tmp / "approved.xlsx")
    imported = read_staging(approved_book, blank_doc)

    out = tmp / "written.pdf"
    written = write_annotations(blank_doc.path, imported.rows, out)
    return Run(annotated=ann_doc, blank=blank_doc, index=index, house=house,
               workbook=approved_book, imported=imported, written=written,
               out_pdf=out)


def approve(book: Path, out: Path) -> Path:
    """Sign off every row pre-fill managed to fill, and nothing else.

    A NEEDS_REVIEW row carries a suggestion but no decision - the columns a
    reviewer fills are left blank on purpose - so approving one means copying
    the suggestion across, which is precisely what a reviewer who agrees with it
    would do. NEEDS_MAPPING rows are left alone: history had nothing to offer
    and there is nothing to agree with.
    """
    wb = load_workbook(book)
    ws = wb[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    col = {n: names.index(n) + 1 for n in names}
    for i in range(2, ws.max_row + 1):
        cell = ws.cell(row=i, column=col["status"])
        if cell.value not in APPROVABLE:
            continue
        for final, suggested in (("final_variable", "suggested_variable"),
                                 ("final_annotation", "suggested_annotation"),
                                 ("final_annot_type", "suggested_annot_type")):
            target = ws.cell(row=i, column=col[final])
            if not target.value:
                target.value = ws.cell(row=i, column=col[suggested]).value
        if ws.cell(row=i, column=col["final_annotation"]).value:
            cell.value = "APPROVED"
    wb.save(out)
    return out
