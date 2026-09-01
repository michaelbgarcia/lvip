"""Fill colour keyed on the statement's SDTM domain, not on its annotation type.

On a real page `DSTERM` and `RFICDTC` are both plain VARIABLE markup, on the
same field, on the same row - and they are drawn in different colours, because
one is DS and the other is DM. `HouseStyle.by_type` cannot express that, so a
domain axis is measured beside it and overrides the type rule for the fill alone.

The hard half is the refusal. DM's own variables carry no prefix, so nothing in
`RFICDTC` says DM; falling back to the *form's* domain would answer DS on the
one page where the difference matters. So the domain is left unresolved, history
answers for the statement if it can, and if it cannot the colour is left blank
and said to be undecided rather than defaulted.
"""
import pytest

from acrf_parser import annotations as ann
from acrf_parser import staging as st, style as sty
from acrf_parser.prefill import PrefillIndex
from acrf_parser.style import derive_house_style


@pytest.fixture(scope="module")
def coloured_house(coloured_doc):
    return derive_house_style(coloured_doc)


# --- reading a statement's domain ------------------------------------------
@pytest.mark.parametrize("text, parsed, expected, how", [
    ("DS=Disposition", {"domain": "DS"}, "DS", ann.QUALIFIED),
    ("DM.BRTHDTC", {"domain": "DM", "variable": "BRTHDTC"}, "DM", ann.QUALIFIED),
    ("DSSTDTC", {"variable": "DSSTDTC"}, "DS", ann.PREFIX),
    ("DSDECOD=INFORMED CONSENT OBTAINED", {"variable": "DSDECOD"}, "DS", ann.PREFIX),
    # DM's own variables carry no prefix - and neither AGE nor SEX may be read as
    # AG- and SE-domain markup on the strength of their first two letters.
    ("RFICDTC", {"variable": "RFICDTC"}, "", ann.UNRESOLVED),
    ("AGE", {"variable": "AGE"}, "", ann.UNRESOLVED),
    ("SEX", {"variable": "SEX"}, "", ann.UNRESOLVED),
])
def test_statement_domain(text, parsed, expected, how):
    assert ann.statement_domain(text, parsed) == (expected, how)


# --- measuring the axis ----------------------------------------------------
def test_the_corpus_is_seen_to_colour_code(coloured_house):
    assert coloured_house.fill_varies_by_domain
    assert coloured_house.by_domain["DS"].fill_color != coloured_house.by_domain["DM"].fill_color


def test_a_single_fill_corpus_keeps_its_one_rule(house):
    """A sponsor that fills every box the same colour must behave exactly as it
    did before this axis existed - one settled rule, not ten under-sampled ones."""
    assert not house.fill_varies_by_domain
    rule, basis = house.for_annotation("VARIABLE", "MHSTDTC")
    assert rule.fill_color == house.for_type("VARIABLE").fill_color
    assert basis == sty.FILL_TYPE


def test_two_statements_of_one_type_take_their_own_domain_s_fill(coloured_house):
    """The whole point: same type, same row, different colour."""
    ds, _ = coloured_house.for_annotation("VARIABLE", "DSSTDTC")
    dm, _ = coloured_house.for_annotation("VARIABLE", "DM.BRTHDTC")
    assert ds.fill_color == coloured_house.by_domain["DS"].fill_color
    assert dm.fill_color == coloured_house.by_domain["DM"].fill_color
    assert ds.fill_color != dm.fill_color
    # Everything but the fill still comes from the type rule.
    assert ds.text_color == dm.text_color and ds.font_size == dm.font_size


def test_an_unprefixed_variable_falls_back_to_its_own_history(coloured_house):
    """RFICDTC. Nothing in the text says DM, so the corpus answers for the exact
    statement instead - which is the only evidence that exists."""
    rule, basis = coloured_house.for_annotation("VARIABLE", "RFICDTC")
    assert rule.fill_color == coloured_house.by_domain["DM"].fill_color
    assert basis == sty.FILL_STATEMENT


def test_an_unknown_statement_is_left_undecided_not_defaulted(coloured_house):
    """No domain and no history. The majority colour is a specific claim about
    which domain this is, and nothing supports it - so nobody makes it."""
    rule, basis = coloured_house.for_annotation("VARIABLE", "USUBJID")
    assert rule.fill_color is None
    assert basis == sty.FILL_UNDECIDED


def test_the_fill_axis_survives_a_round_trip_through_the_database(coloured_doc, tmp_path):
    """The offline path has to reach the same colours as the in-memory one, or
    a workbook built from the corpus would disagree with one built from PDFs."""
    from acrf_parser.kb import KnowledgeBase, build_kb
    from acrf_parser.style import derive_house_style_from_kb
    build_kb(coloured_doc, tmp_path / "corpus.sqlite")
    with KnowledgeBase(tmp_path / "corpus.sqlite") as kb:
        offline = derive_house_style_from_kb(kb)
    live = derive_house_style(coloured_doc)
    assert offline.fill_varies_by_domain == live.fill_varies_by_domain
    assert {k: v.fill_color for k, v in offline.by_domain.items()} == \
           {k: v.fill_color for k, v in live.by_domain.items()}
    assert offline.for_annotation("VARIABLE", "RFICDTC")[0].fill_color == \
           live.for_annotation("VARIABLE", "RFICDTC")[0].fill_color


# --- what the workbook shows ----------------------------------------------
@pytest.fixture(scope="module")
def coloured_rows(coloured_blank, coloured_doc):
    index = PrefillIndex.from_documents([coloured_doc])
    return st.build_staging(coloured_blank, index, derive_house_style(coloured_doc))


def _row(rows, text):
    return next(r for r in rows if r.values["suggested_annotation"] == text)


def test_each_row_carries_the_fill_its_domain_asks_for(coloured_rows):
    assert _row(coloured_rows, "DSTERM").values["fill_rgb"] == "#FFFAC4"
    assert _row(coloured_rows, "RFICDTC").values["fill_rgb"] == "#C7E3F5"


def test_every_row_says_where_its_fill_came_from(coloured_rows):
    """A colour a reviewer cannot account for is the one they should look at."""
    assert "domain DS" in _row(coloured_rows, "DSTERM").values["fill_basis"]
    assert _row(coloured_rows, "RFICDTC").values["fill_basis"] == sty.FILL_STATEMENT


def test_the_palette_sheet_lists_the_domains(coloured_blank, coloured_doc, tmp_path):
    from openpyxl import load_workbook
    path = st.write_staging(coloured_blank, tmp_path / "c.xlsx",
                            index=PrefillIndex.from_documents([coloured_doc]),
                            house=derive_house_style(coloured_doc))
    ws = load_workbook(path)[st.SHEET_FILLS]
    rows = {r[0]: r[1] for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
    assert rows["domain:DS"] == "#FFFAC4" and rows["domain:DM"] == "#C7E3F5"
    assert rows["statement:RFICDTC"] == "#C7E3F5"
