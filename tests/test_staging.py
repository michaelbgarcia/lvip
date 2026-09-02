"""Phase 9 tests: the staging workbook and its contract with the importer."""
import pytest
from openpyxl import load_workbook

from acrf_parser import prefill as pf
from acrf_parser import staging as st
from acrf_parser.normalize import normalize, statement_key
from acrf_parser.prefill import Candidate, PrefillIndex


@pytest.fixture(scope="session")
def book(blank_doc, index, house, tmp_path_factory):
    path = st.write_staging(blank_doc, tmp_path_factory.mktemp("xl") / "staging.xlsx",
                            index=index, house=house)
    return load_workbook(path)


@pytest.fixture(scope="session")
def work(book):
    ws = book[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    return ws, names, [dict(zip(names, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


@pytest.fixture(scope="session")
def multi_index(corpus, blank_doc):
    """History in which "Date of birth" carried two annotations, not one."""
    fld = next(f for f in blank_doc.iter_fields() if f.text == "Date of birth")
    idx = PrefillIndex.from_documents(corpus)
    key = (normalize(fld.form_name), fld.normalized_text)
    idx.key_sets[key][statement_key("RFICDTC")] = Candidate(
        tier=pf.EXACT_KEY, confidence=pf.CONF_EXACT, variable="RFICDTC",
        annotation_text="RFICDTC", annot_type="VARIABLE",
        source="prior.pdf · Demographics · Date of birth",
        evidence=["(form, field_text) seen in prior.pdf"])
    return idx


@pytest.fixture(scope="session")
def multi_book(blank_doc, multi_index, house, tmp_path_factory):
    path = st.write_staging(blank_doc, tmp_path_factory.mktemp("xlm") / "multi.xlsx",
                            index=multi_index, house=house)
    return load_workbook(path)


def _rows(book, sheet=st.SHEET_WORK):
    ws = book[sheet]
    names = [c.value for c in ws[1]]
    return [dict(zip(names, r)) for r in ws.iter_rows(min_row=2, values_only=True)]


# The blank CRF has 14 fields. Form rows are not one per page: they are one per
# *statement* the form carries in its own right, which is the page's domain
# header plus any markup the linker could not place on a field. Page 1 of the
# fixture carries two annotations with no field beside them - deliberately, to
# pin that they are not force-fitted onto a neighbour - and each of those is now
# a row of its own rather than being dropped on the floor.
FIELD_ROWS, FORM_ROWS = 14, 7
ALL_ROWS = FIELD_ROWS + FORM_ROWS
# Pages that belong to a form, and so get an anchor. With no history each is one
# NEEDS_MAPPING row; with history a page's anchor carries a row per statement.
FORM_PAGES = 5
# Form rows history can answer outright. Not all seven: the corpus saw
# "MH=Medical History" on the form's *first* page, so proposing it again on the
# two continuation pages is a suggestion, not a fact, and it arrives as one.
FORM_AUTO = 5


def _field_rows(rows):
    return [r for r in rows if r["scope"] != "FORM"]


def test_sheets(book):
    assert book.sheetnames == [st.SHEET_WORK, st.SHEET_GEOM, st.SHEET_STYLE,
                               st.SHEET_FILLS, st.SHEET_ALTS, st.SHEET_FORMS,
                               st.SHEET_README]


def test_one_row_per_field(work, blank_doc):
    """This corpus has one annotation per field, so rows and fields coincide."""
    _, _, rows = work
    fields = _field_rows(rows)
    assert len(fields) == len(list(blank_doc.iter_fields())) == FIELD_ROWS
    assert [r["row_id"] for r in fields] == [f.id for f in blank_doc.iter_fields()]
    # Of the *fields*. A form row may be one of several on its page's anchor.
    assert {r["annot_seq"] for r in fields} == {1}


def test_auto_rows_arrive_decided(work):
    _, _, rows = work
    auto = [r for r in _field_rows(rows) if r["status"] == "AUTO"]
    assert len(auto) == 9
    assert all(r["match_tier"] == pf.EXACT_KEY for r in auto)
    assert all(r["final_variable"] == r["suggested_variable"] for r in auto)


def test_unresolved_rows_are_left_blank_on_purpose(work):
    """A suggestion must never sit in the answer column looking like a decision."""
    _, _, rows = work
    for r in rows:
        if r["status"] != "AUTO":
            assert not r["final_variable"]
            assert not r["final_annotation"]


def test_evidence_travels_with_the_suggestion(work):
    """The reviewer sees why, next to what - not in a separate audit log."""
    _, names, rows = work
    for col in ("match_tier", "match_score", "match_source", "match_reason"):
        assert col in names
    row = next(r for r in rows if r["row_id"] == "p1f0")
    assert row["match_tier"] == pf.EXACT_KEY and row["match_score"] == 0.95
    assert "sample_acrf.pdf" in row["match_source"]
    assert "seen in" in row["match_reason"]


def test_formatting_is_prefilled_from_house_style(work):
    _, _, rows = work
    row = rows[0]
    assert row["color_rgb"] == "#D91A1A"
    assert row["font_name"] == "Helv" and row["font_size"] == 8.0
    assert row["placement"] == "right_of_field"


def test_status_column_has_a_dropdown(book):
    ws = book[st.SHEET_WORK]
    assert ws.data_validations.dataValidation
    options = " ".join(dv.formula1 for dv in ws.data_validations.dataValidation)
    for status in ("AUTO", "NEEDS_REVIEW", "APPROVED", "REJECTED"):
        assert status in options


def test_evidence_columns_are_locked_and_work_columns_are_not(book):
    ws = book[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    locked = ws.cell(row=2, column=names.index("match_tier") + 1)
    open_ = ws.cell(row=2, column=names.index("final_variable") + 1)
    assert locked.protection.locked and not open_.protection.locked


# --- the importer's half of the contract -----------------------------------
def test_geometry_is_keyed_by_row_id_and_kept_off_the_work_sheet(book, work):
    _, names, rows = work
    geom = book[st.SHEET_GEOM]
    gnames = [c.value for c in geom[1]]
    grows = [dict(zip(gnames, r)) for r in geom.iter_rows(min_row=2, values_only=True)]
    assert [g["row_id"] for g in grows] == [r["row_id"] for r in rows]
    assert "rel_x_pct" not in names          # noise on the work surface
    assert "rel_x_pct" in gnames             # but present for the importer


def test_stored_geometry_is_page_relative(book):
    geom = book[st.SHEET_GEOM]
    gnames = [c.value for c in geom[1]]
    for row in geom.iter_rows(min_row=2, values_only=True):
        g = dict(zip(gnames, row))
        for key in ("rel_x_pct", "rel_y_pct", "rel_w_pct", "rel_h_pct"):
            assert 0.0 <= g[key] <= 1.0
        assert g["page_width"] > 1 and g["page_height"] > 1


def test_house_style_sheet_flags_what_needs_deciding(book):
    ws = book[st.SHEET_STYLE]
    names = [c.value for c in ws[1]]
    rows = [dict(zip(names, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    variable = next(r for r in rows if r["scope"] == "VARIABLE")
    headers = next(r for r in rows if r["scope"] == "DOMAIN_HEADER")
    assert variable["settled"] == "yes"
    assert headers["settled"] == "NO" and headers["size_agreement"] < 1.0


def test_forms_sheet_lists_the_parse(book, blank_doc):
    ws = book[st.SHEET_FORMS]
    rows = [r[0] for r in ws.iter_rows(min_row=2, values_only=True)]
    assert rows == [f.name for f in blank_doc.forms]


# --- several annotations on one field --------------------------------------
def test_a_field_history_saw_twice_gets_two_rows(multi_book):
    rows = [r for r in _rows(multi_book) if r["row_id"] == "p1f0"]
    assert [r["annot_seq"] for r in rows] == [1, 2]
    assert {r["suggested_annotation"] for r in rows} == {"BRTHDTC", "RFICDTC"}
    assert all(r["status"] == "AUTO" for r in rows)


def test_every_other_field_still_gets_exactly_one_row(multi_book, blank_doc):
    counts: dict[str, int] = {}
    for r in _field_rows(_rows(multi_book)):
        counts[r["row_id"]] = counts.get(r["row_id"], 0) + 1
    assert counts.pop("p1f0") == 2
    assert set(counts.values()) == {1}
    assert len(counts) == len(list(blank_doc.iter_fields())) - 1


def test_both_rows_carry_a_position_of_their_own(multi_book):
    """A sibling row gets its own geometry row, and says where it came from.

    Siblings used to share their field's box outright. Where history has no
    position for a statement they still effectively do - and the importer's
    inheritance rule still covers a row a *reviewer* adds, who cannot write to
    the locked Geometry sheet at all. But where history recorded where a
    statement was actually drawn, each sibling carries that spot rather than the
    field's, which is how a set that was spread across a page comes back spread
    instead of piled along one row.
    """
    geom = [g for g in _rows(multi_book, st.SHEET_GEOM) if g["row_id"] == "p1f0"]
    assert [g["annot_seq"] for g in geom] == [1, 2]
    for g in geom:
        assert 0.0 <= g["rel_x_pct"] <= 1.0 and 0.0 <= g["rel_y_pct"] <= 1.0
        assert g["anchor_source"] in (st.LEARNED, st.LEARNED_SPOT, st.HOUSE_STYLE)


def test_the_summary_counts_rows_and_fields_separately(blank_doc, multi_index, house):
    rows = st.build_staging(blank_doc, index=multi_index, house=house)
    s = st.summarize_staging(rows)
    assert s["rows"] == ALL_ROWS + 1 and s["fields"] == FIELD_ROWS
    assert s["form_rows"] == FORM_ROWS
    assert s["multi_annotation_fields"] == 1


def test_instructions_state_the_row_id_rule(book):
    text = "\n".join(str(r[0]) for r in book[st.SHEET_README].iter_rows(values_only=True))
    assert "row_id" in text and "Do not sort or reorder rows" in text
    assert "NEEDS_MAPPING" in text


def test_instructions_explain_how_to_add_a_second_annotation(book):
    """The one thing a reviewer cannot work out from the columns alone."""
    text = "\n".join(str(r[0]) for r in book[st.SHEET_README].iter_rows(values_only=True))
    assert "Adding a second annotation to a field" in text
    assert "annot_seq" in text and "Keep row_id exactly as it is" in text


# --- no history yet --------------------------------------------------------
def test_a_first_study_still_produces_a_workbook(blank_doc, tmp_path):
    """No corpus is not an error - it is a sheet where every row is the agent's."""
    path = st.write_staging(blank_doc, tmp_path / "cold.xlsx", index=PrefillIndex())
    ws = load_workbook(path)[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    rows = [dict(zip(names, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    # One row per field, and one per page that belongs to a form. The extra form
    # rows only exist where history has statements to put in them, and here
    # there is no history - so this is the floor, not ALL_ROWS.
    assert len(rows) == FIELD_ROWS + FORM_PAGES
    assert all(r["status"] == "NEEDS_MAPPING" for r in rows)
    assert all(not r["final_variable"] for r in rows)


def test_summary(blank_doc, index, house):
    s = st.summarize_staging(st.build_staging(blank_doc, index, house))
    assert s["rows"] == ALL_ROWS and s["fields"] == FIELD_ROWS
    assert s["by_status"]["AUTO"] == 9 + FORM_AUTO


# --- the Alternatives sheet ------------------------------------------------
@pytest.fixture(scope="session")
def rival_book(blank_doc, corpus, house, tmp_path_factory):
    """A corpus where a second study worded the same markup differently."""
    import json

    idx = PrefillIndex.from_documents(corpus)
    rows = []
    for key, statements in list(idx.form_sets.items()):
        for c in statements.values():
            rows.append({"file_name": "rival.pdf", "form_name": key,
                         "normalized_name": key, "domain": "",
                         "annotation_text": c.annotation_text + " Status",
                         "annot_type": c.annot_type, "variable": c.variable,
                         "annotation_bbox": [40, 40, 120, 52],
                         "annotation_relative": json.dumps(
                             {"rel_x_pct": 0.3, "rel_y_pct": 0.05,
                              "rel_w_pct": 0.13, "rel_h_pct": 0.015}),
                         "page_offset": c.drawn.get("page_offset"),
                         "trust": pf.GEOMETRIC})
    idx._load_form_annotations(rows)
    path = st.write_staging(blank_doc, tmp_path_factory.mktemp("xlr") / "rival.xlsx",
                            index=idx, house=house)
    return load_workbook(path)


def test_the_alternatives_sheet_exists_even_when_the_corpus_agrees(book):
    """Header row always, so the agent's instructions never point at nothing."""
    ws = book[st.SHEET_ALTS]
    assert [c.value for c in ws[1]] == [h for h, _ in st.ALT_HEADERS]


def test_rival_wordings_land_on_the_alternatives_sheet(rival_book):
    alts = _rows(rival_book, st.SHEET_ALTS)
    assert alts, "a disagreeing corpus must produce alternatives"
    assert all(a["alt_annotation"] and a["alt_source"] for a in alts)
    # Which side wins is the rule's business; that the disagreement is on the
    # sheet is this test's. One of the two wordings is the rival.
    assert any(a["alt_annotation"].endswith("Status")
               or a["chosen_annotation"].endswith("Status") for a in alts)


def test_alternatives_never_become_annotation_rows(book, rival_book):
    """The property that matters: a rival must not be drawable. Same CRF, more
    history, and not one extra row on the sheet the importer reads."""
    assert len(_rows(rival_book)) == len(_rows(book)) == ALL_ROWS


def test_an_alternative_is_keyed_back_to_the_row_it_rivals(rival_book):
    work = {r["row_id"] for r in _rows(rival_book)}
    alts = _rows(rival_book, st.SHEET_ALTS)
    assert {a["row_id"] for a in alts} <= work
    for a in alts:
        assert a["chosen_annotation"] != a["alt_annotation"]


def test_a_row_with_alternatives_says_so_in_its_evidence(rival_book):
    """The pointer, so the agent knows which rows to go and look up."""
    ids = {a["row_id"] for a in _rows(rival_book, st.SHEET_ALTS)}
    noted = {r["row_id"] for r in _rows(rival_book)
             if "Alternatives" in (r["match_reason"] or "")}
    assert ids <= noted


def test_the_alternatives_sheet_is_locked(rival_book):
    ws = rival_book[st.SHEET_ALTS]
    assert all(c.protection.locked for c in ws[2])


def test_the_instructions_tell_the_agent_what_to_do_with_them(book):
    text = "\n".join(str(r[0].value or "") for r in book[st.SHEET_README].iter_rows())
    assert "Reviewing the 'Alternatives' sheet" in text
    assert "REPLACES the" in text and "final_annotation" in text
