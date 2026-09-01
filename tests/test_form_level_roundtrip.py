"""The whole loop for the form-level layer, checked by re-reading the PDF.

Every other test here can pass while the output PDF is still missing the markup
across the top of its pages, because the failure was never an error - the rows
simply did not exist. So the assertion that matters is made against the written
file, parsed fresh by the same parser, with no workbook to consult.
"""
import pytest
from openpyxl import load_workbook

from acrf_parser import parse_pdf, staging as st
from acrf_parser.importer import read_staging
from acrf_parser.kb import KnowledgeBase, ingest_approved
from acrf_parser.prefill import PrefillIndex
from acrf_parser.style import derive_house_style
from acrf_parser.writer import write_annotations


@pytest.fixture
def approved(coloured_blank, coloured_doc, tmp_path):
    """A staging workbook pre-filled from history and signed off wholesale."""
    index = PrefillIndex.from_documents([coloured_doc])
    path = st.write_staging(coloured_blank, tmp_path / "c.xlsx", index=index,
                            house=derive_house_style(coloured_doc))
    wb = load_workbook(path)
    ws = wb[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    status = names.index("status") + 1
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=status).value == "AUTO":
            ws.cell(row=i, column=status).value = "APPROVED"
    out = tmp_path / "returned.xlsx"
    wb.save(out)
    return read_staging(out, coloured_blank)


# --- the sheet -------------------------------------------------------------
def test_the_page_s_form_level_markup_reaches_the_sheet(coloured_blank, coloured_doc):
    """Three statements on one page, each its own row, in the drawn order."""
    index = PrefillIndex.from_documents([coloured_doc])
    rows = st.build_staging(coloured_blank, index, derive_house_style(coloured_doc))
    page1 = [r for r in rows if r.row_id == "p1h"]
    assert [r.values["suggested_annotation"] for r in page1] == [
        "DS=Disposition", "DM=Demographics", "DSCAT = PROTOCOL MILESTONE"]
    assert [r.annot_seq for r in page1] == [1, 2, 3]
    assert {r.values["scope"] for r in page1} == {"FORM"}


def test_the_form_rows_lead_their_page(coloured_blank, coloured_doc):
    """Reading order: what is printed across the top of the page comes first."""
    rows = st.build_staging(coloured_blank, PrefillIndex.from_documents([coloured_doc]),
                            derive_house_style(coloured_doc))
    assert rows[0].row_id == "p1h"
    assert rows[0].values["field_text"] == "[form header] Informed Consent"


def test_a_first_study_is_asked_about_the_form_layer(coloured_blank):
    """With no history there is still a row per page, unfilled. Being asked is
    the fix: before, the layer was absent from the sheet and nobody was."""
    rows = st.build_staging(coloured_blank, PrefillIndex())
    form_rows = [r for r in rows if r.values["scope"] == "FORM"]
    assert len(form_rows) == 2
    assert all(r.values["status"] == "NEEDS_MAPPING" for r in form_rows)
    assert "name the form's domain(s) here" in form_rows[0].values["match_reason"]


# --- the import ------------------------------------------------------------
def test_form_rows_import_and_are_approved(approved):
    assert not approved.errors
    assert len(approved.form_rows()) == 4          # three on page 1, one on page 2
    assert all(r.ready for r in approved.form_rows())


def test_a_page_may_legitimately_carry_no_form_markup(coloured_blank, coloured_doc,
                                                      tmp_path):
    """A field with no rows is an error; a page with none is ordinary."""
    index = PrefillIndex.from_documents([coloured_doc])
    path = st.write_staging(coloured_blank, tmp_path / "d.xlsx", index=index)
    wb = load_workbook(path)
    ws = wb[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    for i in reversed(range(2, ws.max_row + 1)):
        if ws.cell(row=i, column=names.index("row_id") + 1).value == "p2h":
            ws.delete_rows(i)
    out = tmp_path / "nop2h.xlsx"
    wb.save(out)
    report = read_staging(out, coloured_blank)
    assert not [i for i in report.issues if i.code == "MISSING_ROW"]


def test_the_scope_column_cannot_move_a_row_onto_a_field(coloured_blank, coloured_doc,
                                                         tmp_path):
    """row_id decides what a row is; the column is checked against it, not believed."""
    index = PrefillIndex.from_documents([coloured_doc])
    path = st.write_staging(coloured_blank, tmp_path / "e.xlsx", index=index)
    wb = load_workbook(path)
    ws = wb[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    scope = names.index("scope") + 1
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=names.index("row_id") + 1).value == "p1h":
            ws.cell(row=i, column=scope).value = "FIELD"
            break
    out = tmp_path / "mixed.xlsx"
    wb.save(out)
    report = read_staging(out, coloured_blank)
    row = report.rows_for("p1h")[0]
    assert row.scope == "FORM"
    assert "SCOPE_MISMATCH" in {i.code for i in row.issues}
    assert not [i for i in row.issues if i.severity == "ERROR"]


# --- the PDF ---------------------------------------------------------------
def test_the_headers_are_drawn_across_the_top_of_the_page(approved, coloured_blank,
                                                          tmp_path):
    out = tmp_path / "annotated.pdf"
    report = write_annotations(coloured_blank.path, approved.rows, out)
    assert report.to_dict()["form_annotations"] == 4

    again = parse_pdf(out)
    # The three form-level statements, in the order they were drawn. Read off
    # the *header band* rather than off `form_annotations`: that now also holds
    # field markup the linker could not place on the re-parse, which belongs to
    # the form for want of anywhere better and is a different claim from "this
    # is the row of headers across the top".
    wanted = ["DS=Disposition", "DM=Demographics", "DSCAT = PROTOCOL MILESTONE"]
    boxes = {a.text: a.bbox for a in again.form_annotations(1)}
    assert set(wanted) <= set(boxes)
    band = [boxes[t] for t in wanted]
    # Above every field on the page, and left to right along one band.
    assert max(b.y1 for b in band) <= min(f.bbox.y0 for f in again.page(1).fields)
    assert [b.x0 for b in band] == sorted(b.x0 for b in band)
    assert len({round(b.y0, 1) for b in band}) == 1


def test_the_domain_fills_land_on_the_page(approved, coloured_blank, tmp_path):
    """Two statements of one type on one row, in two colours, as drawn."""
    out = tmp_path / "filled.pdf"
    write_annotations(coloured_blank.path, approved.rows, out)
    # Page 1 only: both pages carry a DM=Demographics header, and collapsing
    # them by text would quietly compare the wrong one.
    page1 = {a.text: a for a in parse_pdf(out).iter_annotations() if a.page == 1}
    assert page1["DSTERM"].fill_color != page1["RFICDTC"].fill_color
    assert page1["DSSTDTC"].fill_color == page1["DSTERM"].fill_color
    # The header takes its domain's colour like anything else, so the DM header
    # and the DM variable under it match while the DS header beside it does not.
    assert page1["DM=Demographics"].fill_color == page1["RFICDTC"].fill_color
    assert page1["DS=Disposition"].fill_color == page1["DSTERM"].fill_color


def test_the_form_layer_is_learned_back_into_the_corpus(approved, coloured_blank,
                                                        tmp_path):
    """Without this the next study starts from nothing again - and a layer that
    is re-solved from scratch every time is a layer that will be forgotten."""
    db = tmp_path / "learned.sqlite"
    ingest_approved(approved, coloured_blank, db, source="reviewed")
    with KnowledgeBase(db) as kb:
        rows = [dict(r) for r in kb.con.execute(
            "SELECT * FROM form_annotations ORDER BY annotation_bbox")]
        index = PrefillIndex.from_kb(kb)
    assert {r["annotation_text"] for r in rows} == {
        "DS=Disposition", "DM=Demographics", "DSCAT = PROTOCOL MILESTONE"}
    assert all(r["trust"] == "HUMAN_APPROVED" for r in rows)

    # And it comes back as the same set, in the same order, for the same form.
    again = st.build_staging(coloured_blank, index)
    page1 = [r for r in again if r.row_id == "p1h"]
    assert [r.values["suggested_annotation"] for r in page1] == [
        "DS=Disposition", "DM=Demographics", "DSCAT = PROTOCOL MILESTONE"]
    assert {r.values["status"] for r in page1} == {"AUTO"}
