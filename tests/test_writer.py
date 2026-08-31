"""Phase 10 tests: drawing annotations, and the loop closing on itself."""
from pathlib import Path

import pymupdf
import pytest

from acrf_parser import parse_pdf
from acrf_parser.importer import ERROR, ImportedRow, Issue, read_staging
from acrf_parser.writer import write_annotations
from acrf_parser import staging as st
from openpyxl import load_workbook


def _row(row_id, rel_y, text="MHSTDTC", rel_x=0.4, status="APPROVED",
         placement="right_of_field", size=8.0, page_w=595.0):
    return ImportedRow(
        row_id=row_id, form_name="Medical History", field_text=row_id, status=status,
        annotation_text=text, annot_type="VARIABLE", text_color=(0.85, 0.1, 0.1),
        font_name="Helv", font_size=size, placement=placement,
        geometry={"page": 1, "page_width": page_w, "page_height": 842.0,
                  "rel_x_pct": rel_x, "rel_y_pct": rel_y,
                  "rel_w_pct": 0.08, "rel_h_pct": 0.016,
                  "offset_x_pct": 0.02, "offset_y_pct": 0.0})


@pytest.fixture(scope="session")
def approved(blank_doc, index, house, tmp_path_factory):
    """The blank CRF, staged and signed off - what the writer is meant to consume."""
    d = tmp_path_factory.mktemp("approve")
    path = st.write_staging(blank_doc, d / "s.xlsx", index=index, house=house)
    wb = load_workbook(path)
    ws = wb[st.SHEET_WORK]
    names = [c.value for c in ws[1]]
    col = names.index("status") + 1
    for i in range(2, ws.max_row + 1):
        if ws.cell(row=i, column=col).value == "AUTO":
            ws.cell(row=i, column=col).value = "APPROVED"
    out = d / "approved.xlsx"
    wb.save(out)
    return read_staging(out, blank_doc)


# --- the loop --------------------------------------------------------------
def test_written_annotations_survive_a_reparse(approved, blank_doc, tmp_path):
    """The whole pipeline, checked by an independent read of its own output.

    Write the approved rows onto the blank CRF, then parse the result with the
    same parser - form detection, field extraction, classification and linking
    all running fresh - and require every mapping the sheet specified to come
    back attached to the field it was specified for.
    """
    out = tmp_path / "annotated.pdf"
    report = write_annotations(blank_doc.path, approved.rows, out)
    assert len(report.placements) == 9 and not report.adjusted

    again = parse_pdf(out)
    assert len(list(again.iter_annotations())) == 9

    linked = {l.field_id: again.annotation(l.annotation_id).text
              for l in again.links if not l.rejected}
    expected = {r.row_id: r.text_to_draw for r in approved.approved()}
    assert linked == expected


def test_the_style_the_sheet_asked_for_is_what_lands(approved, blank_doc, tmp_path):
    out = tmp_path / "styled.pdf"
    write_annotations(blank_doc.path, approved.rows, out)
    a = next(a for a in parse_pdf(out).iter_annotations() if a.text == "BRTHDTC")
    assert a.text_color == (0.851, 0.102, 0.102)     # #D91A1A, 8-bit round trip
    assert a.font_name == "Helv" and a.font_size == 8.0


def test_written_annotations_are_stripped_from_the_text_layer(approved, blank_doc, tmp_path):
    """Phase 1's from_annotation flag must hold for annotations we wrote too,
    or a re-parse would read our own markup back as CRF field labels."""
    out = tmp_path / "annotated.pdf"
    write_annotations(blank_doc.path, approved.rows, out)
    again = parse_pdf(out)
    labels = {f.text for f in again.iter_fields()}
    assert not labels & {"BRTHDTC", "MHSTDTC", "MHENDTC"}
    assert "Date of birth" in labels


# --- what gets drawn -------------------------------------------------------
def test_only_approved_and_valid_rows_are_drawn(blank_doc, tmp_path):
    rows = [_row("ok", 0.2),
            _row("pending", 0.3, status="NEEDS_MAPPING"),
            _row("reviewed", 0.4, status="NEEDS_REVIEW")]
    blocked = _row("blocked", 0.5)
    blocked.issues.append(Issue("blocked", "color_rgb", ERROR, "BAD_COLOR",
                                "'crimson' is not a #RRGGBB colour"))
    rows.append(blocked)

    report = write_annotations(blank_doc.path, rows, tmp_path / "part.pdf")
    assert [p.row_id for p in report.placements] == ["ok"]
    assert dict(report.skipped) == {
        "pending": "status is NEEDS_MAPPING, not APPROVED",
        "reviewed": "status is NEEDS_REVIEW, not APPROVED",
        "blocked": "blocked by validation"}


def test_nothing_is_written_in_place(approved, blank_doc, tmp_path):
    before = Path(blank_doc.path).read_bytes()
    write_annotations(blank_doc.path, approved.rows, tmp_path / "copy.pdf")
    assert Path(blank_doc.path).read_bytes() == before


# --- placement -------------------------------------------------------------
def test_annotation_is_placed_right_of_its_field(blank_doc, tmp_path):
    report = write_annotations(blank_doc.path, [_row("a", 0.3)], tmp_path / "p.pdf")
    rect = report.placements[0].rect
    field_x1 = (0.4 + 0.08 / 2) * 595
    assert rect.x0 > field_x1                       # to the right
    assert abs(rect.cy - 0.3 * 842) < 2             # on the field's row


@pytest.mark.parametrize("placement,check", [
    ("left_of_field", lambda r, f: r.x1 <= f["x0"]),
    ("below_field", lambda r, f: r.y0 >= f["y1"]),
    ("above_field", lambda r, f: r.y1 <= f["y0"]),
])
def test_placement_labels_are_honoured(blank_doc, tmp_path, placement, check):
    report = write_annotations(blank_doc.path, [_row("a", 0.5, placement=placement)],
                               tmp_path / f"{placement}.pdf")
    rect = report.placements[0].rect
    field = {"x0": (0.4 - 0.04) * 595, "x1": (0.4 + 0.04) * 595,
             "y0": (0.5 - 0.008) * 842, "y1": (0.5 + 0.008) * 842}
    assert check(rect, field)


def test_a_box_that_runs_off_the_page_is_pulled_back(blank_doc, tmp_path):
    """An annotation half off the page is worse than one a reviewer was told about."""
    long_text = 'SUPPDM.QVAL when QNAM = "VERYLONGQUALIFIERNAME"'
    report = write_annotations(blank_doc.path, [_row("wide", 0.2, long_text, rel_x=0.85)],
                               tmp_path / "wide.pdf")
    p = report.placements[0]
    assert p.rect.x1 <= 595
    assert any("stay on the page" in a for a in p.adjustments)
    assert report.adjusted == [p]


def test_collisions_are_nudged_apart_and_reported(blank_doc, tmp_path):
    """Three fields a fraction of a row apart would otherwise overprint."""
    rows = [_row("a", 0.500), _row("b", 0.502), _row("c", 0.504)]
    report = write_annotations(blank_doc.path, rows, tmp_path / "hit.pdf")
    rects = [p.rect for p in report.placements]
    pairs = [(i, j) for i, a in enumerate(rects) for j, b in enumerate(rects[i + 1:], i + 1)
             if a.h_overlap(b) > 0 and a.v_overlap(b) > 0]
    assert pairs == []
    assert len(report.adjusted) == 2
    assert all("clear another annotation" in "; ".join(p.adjustments)
               for p in report.adjusted)


def test_markup_is_not_printed_over_the_forms_own_text(blank_doc, tmp_path):
    """The anchor the house style computes is frequently already occupied.

    A field near the left margin puts "right of the label, 12pt across" straight
    on top of the label beside it - which is the one thing markup must never do,
    since the reader needs both.
    """
    row = _row("over", 180 / 842, rel_x=0.02)      # anchored onto the "Sex" label
    report = write_annotations(blank_doc.path, [row], tmp_path / "clear.pdf")
    rect = report.placements[0].rect

    words = [w[:4] for w in pymupdf.open(blank_doc.path)[0].get_text("words")]
    assert not [w for w in words
                if min(rect.x1, w[2]) - max(rect.x0, w[0]) > 0.1
                and min(rect.y1, w[3]) - max(rect.y0, w[1]) > 0.1]
    assert any("clear the form text" in a for a in report.placements[0].adjustments)


def test_the_same_statement_is_drawn_once_per_row(blank_doc, tmp_path):
    """A question and its options pre-fill to one variable; the row states it once.

    The two texts here are the same mapping written two ways, which is how it
    actually arrives from a reviewer working row by row. The option sits a shade
    higher than its question, as options do - the question still keeps the
    markup, because on a row it is the leftmost field that owns it.
    """
    rows = [_row("question", 0.300, "SUPPDM.QVAL when QNAM=RACEOR", rel_x=0.15),
            _row("option", 0.298, 'QVAL when SUPPDM.QNAM = "RACEOR"', rel_x=0.60)]
    report = write_annotations(blank_doc.path, rows, tmp_path / "dup.pdf")
    assert [p.row_id for p in report.placements] == ["question"]
    assert dict(report.skipped)["option"] == (
        "same statement as question, already placed on this row")


def test_a_repeat_down_a_column_is_still_drawn(blank_doc, tmp_path):
    """Only side-by-side repeats are one statement. A log form's rows are not."""
    rows = [_row("r1", 0.30), _row("r2", 0.36), _row("r3", 0.42)]
    report = write_annotations(blank_doc.path, rows, tmp_path / "log.pdf")
    assert [p.row_id for p in report.placements] == ["r1", "r2", "r3"]
    assert not report.skipped


def test_an_unavailable_font_is_substituted_and_said_so(blank_doc, tmp_path):
    row = _row("f", 0.3)
    row.font_name = "SponsorSans"
    report = write_annotations(blank_doc.path, [row], tmp_path / "font.pdf")
    assert "font substituted for SponsorSans" in report.placements[0].adjustments
    assert parse_pdf(tmp_path / "font.pdf").page(1).annotations[0].text == "MHSTDTC"


def test_a_row_pointing_off_the_document_is_skipped(blank_doc, tmp_path):
    row = _row("ghost", 0.3)
    row.geometry["page"] = 99
    report = write_annotations(blank_doc.path, [row], tmp_path / "ghost.pdf")
    assert not report.placements
    assert "not in" in dict(report.skipped)["ghost"]


def test_variable_alone_is_enough_to_draw(blank_doc, tmp_path):
    """A reviewer who fills only final_variable still gets an annotation."""
    row = _row("v", 0.3, text="")
    row.variable = "MHTERM"
    report = write_annotations(blank_doc.path, [row], tmp_path / "v.pdf")
    assert report.placements[0].text == "MHTERM"


def test_report_summary(approved, blank_doc, tmp_path):
    report = write_annotations(blank_doc.path, approved.rows, tmp_path / "s.pdf")
    d = report.to_dict()
    assert d["written"] == 9 and d["skipped"] == 5 and d["adjusted"] == 0
